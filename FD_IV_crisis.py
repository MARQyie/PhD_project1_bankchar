#------------------------------------------
# IV models pre, during, post crisis size for first working paper
# Mark van der Plaat
# October 2019 

 # Import packages
import pandas as pd
import numpy as np
import scipy.stats as sps # used to calculated cdf and pdf

import os
os.chdir(r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char')

# Import method that adds a constant to a df
from statsmodels.tools.tools import add_constant

# Import method that can estimate a pooled probit
from statsmodels.discrete.discrete_model import Probit

# Import method for POLS (also does FE)
from linearmodels import PanelOLS

# Import packages for the Sargan-Hausman test
from linearmodels.iv._utility import annihilate
from linearmodels.utility import WaldTestStatistic

import sys # to use the help functions needed
sys.path.insert(1, r'X:\My Documents\PhD\Coding_docs\Help_functions')

from summary3 import summary_col

#--------------------------------------------
# Set parameters 
log = False # If set to False the program estimates the model without logs and with size

#---------------------------------------------- 
#----------------------------------------------
# Prelims
#----------------------------------------------
#----------------------------------------------

#----------------------------------------------
# Load data and add needed variables

# Load df
df = pd.read_csv('df_wp1_clean.csv', index_col = 0)

## Make multi index
df.date = pd.to_datetime(df.date.astype(str).str.strip() + '1230')
df.set_index(['IDRSSD','date'],inplace=True)

## Drop missings on distance
df.dropna(subset = ['distance'], inplace = True)

## Dummy variable for loan sales
if log:
    df['dum_ls'] = np.exp((df.ls_tot > 0) * 1) - 1 #will be taken the log of later
else:
    df['dum_ls'] = (df.ls_tot > 0) * 1  

## Take a subset of variables (only the ones needed)
vars_needed = ['distance','provratio','rwata','net_coffratio_tot_ta',\
               'allowratio_tot_ta','ls_tot_ta','dum_ls','size',\
               'RC7205','loanratio','roa','depratio','comloanratio','RC2170',\
               'num_branch', 'bhc', 'RIAD4150', 'perc_limited_branch',\
               'perc_full_branch', 'unique_states','UNIT']
df_sub = df[vars_needed]

## drop NaNs
df_sub.dropna(subset = ['provratio','rwata','net_coffratio_tot_ta','allowratio_tot_ta',\
               'ls_tot_ta','RC7205','loanratio','roa',\
               'depratio','comloanratio','RC2170','size'], inplace = True)

#---------------------------------------------------
# Setup the data

## Set aside TA
ta = df_sub.RC2170

## Take logs of the df
if log:
    df_sub = df_sub.transform(lambda df: np.log(1 + df)).replace([np.inf, -np.inf], 0)

## Add TA for subsetting
df_sub['ta_sub'] = ta

## Add the x_xbar to the df
if log:
    x_xbar = df_sub[['RC7205','loanratio','roa',\
                      'depratio','comloanratio','RC2170','bhc']].transform(lambda df: df - df.mean())
    df_sub[[x + '_xbar' for x in ['RC7205','loanratio','roa',\
                      'depratio','comloanratio','size','bhc']]] = x_xbar
else:
    x_xbar = df_sub[['RC7205','loanratio','roa',\
                      'depratio','comloanratio','size','bhc']].transform(lambda df: df - df.mean())
    df_sub[[x + '_xbar' for x in ['RC7205','loanratio','roa',\
                      'depratio','comloanratio','size','bhc']]] = x_xbar
# Subset the df
''' Crisis dates are:
        Pre-crisis: 2001-2006
        Crisis: 2007-2009
        Post-crisis: 2010-2018
    '''

df_pre = df_sub[df_sub.index.get_level_values(1) <= pd.Timestamp(2006,12,30)]
df_during = df_sub[(df_sub.index.get_level_values(1) > pd.Timestamp(2006,12,30)) & (df_sub.index.get_level_values(1) < pd.Timestamp(2010,12,30))]
df_post = df_sub[df_sub.index.get_level_values(1) >= pd.Timestamp(2010,12,30)]

## Take the first differences
df_pre_fd = df_pre.groupby(df_pre.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()
df_during_fd = df_during.groupby(df_during.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()
df_post_fd = df_post.groupby(df_post.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()

## Add dummies
dummy_pre_fd = pd.get_dummies(df_pre_fd.index.get_level_values(1))
dummy_during_fd = pd.get_dummies(df_during_fd.index.get_level_values(1))
dummy_post_fd = pd.get_dummies(df_post_fd.index.get_level_values(1))

### Add dummies to the dfs
col_dummy_pre = ['dum' + dummy for dummy in dummy_pre_fd.columns.astype(str).str[:4].tolist()]
col_dummy_during = ['dum' + dummy for dummy in dummy_during_fd.columns.astype(str).str[:4].tolist()]
col_dummy_post = ['dum' + dummy for dummy in dummy_post_fd.columns.astype(str).str[:4].tolist()]
dummy_pre_fd = pd.DataFrame(np.array(dummy_pre_fd), index = df_pre_fd.index, columns = col_dummy_pre)
dummy_during_fd = pd.DataFrame(np.array(dummy_during_fd), index = df_during_fd.index, columns = col_dummy_during)
dummy_post_fd = pd.DataFrame(np.array(dummy_post_fd), index = df_post_fd.index, columns = col_dummy_post)

df_pre_fd[col_dummy_pre] = dummy_pre_fd
df_during_fd[col_dummy_during] = dummy_during_fd
df_post_fd[col_dummy_post] = dummy_post_fd

#---------------------------------------------------
# Load the necessary functions

def fTestWeakInstruments(y, fitted_full, fitted_reduced, dof = 4):
    ''' Simple F-test to test the strength of instrumental variables.
        
        y : True y values
        fitted_full : fitted values of the first stage with instruments
        fitted_reduced : fitted values first stage without instruments
        dof : number of instruments'''
    
    # Calculate the SSE and MSE
    sse_full = np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    sse_reduced =  np.sum([(y.values[i] - fitted_reduced.values[i][0])**2 for i in range(y.shape[0])])
    
    mse_full = (1 / y.shape[0]) * np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    
    # Calculate the statistic
    f_stat = ((sse_reduced - sse_full)/dof) / mse_full
    
    return f_stat

def sargan(resids, x, z, nendog = 1):
    '''Function performs a sargan test (no heteroskedasity) to check
        the validity of the overidentification restrictions. H0: 
        overidentifying restrictions hold.
        
        resids : residuals of the second stage
        x : exogenous variables
        z : instruments 
        nendog : number of endogenous variables'''  

    nobs, ninstr = z.shape
    name = 'Sargan\'s test of overidentification'

    eps = resids.values[:,None]
    u = annihilate(eps, pd.concat([x,z],axis = 1))
    stat = nobs * (1 - (u.T @ u) / (eps.T @ eps)).squeeze()
    null = 'The overidentification restrictions are valid'

    return WaldTestStatistic(stat, null, ninstr - nendog, name=name)

#----------------------------------------------
#----------------------------------------------
# Analyses
#----------------------------------------------
#----------------------------------------------

if log:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + RC2170 + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + RC2170_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + RC2170_G_hat_ls + bhc_G_hat_ls'
else:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + size + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + size_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + size_G_hat_ls + bhc_G_hat_ls' 
    
righthand_z = r'unique_states + RIAD4150'
num_z = righthand_z.count('+') + 1 # Use string count as a trick to count the number of vars in z

time_dummies_pre = ' + '.join(col_dummy_pre[1:])
time_dummies_during = ' + '.join(col_dummy_during[1:])
time_dummies_post = ' + '.join(col_dummy_post[1:])

#----------------------------------------------
# MODEL 1: Pre Crisis
#----------------------------------------------

# First check the data on column rank
if log:
    rank_pre = np.linalg.matrix_rank(df_pre_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_pre = np.linalg.matrix_rank(df_pre_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)

#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat
                
## Dummy LS
res_pre_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step1_w.summary)
df_pre_fd['G_hat_fd_w'] = res_pre_step1_w.fitted_values

## LS/TA
res_pre_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step1_ls.summary)
df_pre_fd['G_hat_fd_ls'] = res_pre_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_pre_fd.loc[:,df_pre_fd.columns.str.contains('_xbar')] * df_pre_fd.G_hat_fd_w[:, None]
if log:
    df_pre_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_pre_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_pre_fd.loc[:,df_pre_fd.columns.str.contains('_xbar')] * df_pre_fd.G_hat_fd_ls[:, None]
if log:
    df_pre_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_pre_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls
    
#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_pre_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_charge_w.summary)

res_pre_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_pre_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_allow_w.summary)

res_pre_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_pre_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_rwata_w.summary)

res_pre_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_pre_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_prov_w.summary)

res_pre_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_pre_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_charge_ls.summary)

res_pre_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_pre_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_allow_ls.summary)

res_pre_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_pre_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_rwata_ls.summary)

res_pre_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_pre_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2a_prov_ls.summary)

res_pre_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_pre_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_pre_step1b_w = fTestWeakInstruments(df_pre_fd.dum_ls, res_pre_step1_w.fitted_values, res_pre_step1b_w.fitted_values, 2) 

### LS/TA
res_pre_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_pre_step1b_ls = fTestWeakInstruments(df_pre_fd.ls_tot_ta, res_pre_step1_ls.fitted_values, res_pre_step1b_ls.fitted_values, 2)

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_pre_fd['resid_step1_w'] = res_pre_step1_w.resids

#----------------------------------------------
#### Charge-off
res_pre_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_charge_w_endo.summary) 

#----------------------------------------------
#### Allow
res_pre_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_allow_w_endo.summary) 

#----------------------------------------------
#### rwata
res_pre_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_rwata_w_endo.summary)

#----------------------------------------------
#### prov
res_pre_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_prov_w_endo.summary) 

#----------------------------------------------
### LS/TA
df_pre_fd['resid_step1_ls'] = res_pre_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_pre_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_charge_ls_endo.summary)

#----------------------------------------------
#### Allow
res_pre_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_allow_ls_endo.summary) 

#----------------------------------------------
#### rwata
res_pre_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_rwata_ls_endo.summary) 

#----------------------------------------------
#### prov
res_pre_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_pre, data = df_pre_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_pre_prov_ls_endo.summary) 

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_pre_step2a_charge_w = sargan(res_pre_step2a_charge_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0000
oir_pre_step2b_charge_w = sargan(res_pre_step2b_charge_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### Allow
oir_pre_step2a_allow_w = sargan(res_pre_step2a_allow_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0000
oir_pre_step2b_allow_w = sargan(res_pre_step2b_allow_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### rwata
oir_pre_step2a_rwata_w = sargan(res_pre_step2a_rwata_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0023
oir_pre_step2b_rwata_w = sargan(res_pre_step2b_rwata_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0015

#----------------------------------------------
#### prov
oir_pre_step2a_prov_w = sargan(res_pre_step2a_prov_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.6458
oir_pre_step2b_prov_w = sargan(res_pre_step2b_prov_w.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.5889

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_pre_step2a_charge_ls = sargan(res_pre_step2a_charge_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0049
oir_pre_step2b_charge_ls = sargan(res_pre_step2b_charge_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0063

#----------------------------------------------
#### Allow
oir_pre_step2a_allow_ls = sargan(res_pre_step2a_allow_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0058
oir_pre_step2b_allow_ls = sargan(res_pre_step2b_allow_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0069

#----------------------------------------------
#### rwata
oir_pre_step2a_rwata_ls = sargan(res_pre_step2a_rwata_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0027
oir_pre_step2b_rwata_ls = sargan(res_pre_step2b_rwata_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.0030

#----------------------------------------------
#### prov
oir_pre_step2a_prov_ls = sargan(res_pre_step2a_prov_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.8760
oir_pre_step2b_prov_ls = sargan(res_pre_step2b_prov_ls.resids, df_pre_fd[righthand_x.split(' + ')], df_pre_fd[righthand_z.split(' + ')]) # p-val = 0.9156


'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 2: During Crisis
#----------------------------------------------
'''-----------------------------------------''' 

if log:
    rank_during = np.linalg.matrix_rank(df_during_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_during = np.linalg.matrix_rank(df_during_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
          
#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat
                
## Dummy LS
res_during_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step1_w.summary)
df_during_fd['G_hat_fd_w'] = res_during_step1_w.fitted_values

## LS/TA
res_during_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step1_ls.summary)
df_during_fd['G_hat_fd_ls'] = res_during_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_during_fd.loc[:,df_during_fd.columns.str.contains('_xbar')] * df_during_fd.G_hat_fd_w[:, None]
if log:
    df_during_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_during_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_during_fd.loc[:,df_during_fd.columns.str.contains('_xbar')] * df_during_fd.G_hat_fd_ls[:, None]
if log:
    df_during_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_during_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls

#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_during_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_charge_w.summary)

res_during_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_during_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_allow_w.summary)

res_during_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_during_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_rwata_w.summary)

res_during_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_during_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_prov_w.summary)

res_during_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_during_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_charge_ls.summary)

res_during_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_during_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_allow_ls.summary)

res_during_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_during_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_rwata_ls.summary)

res_during_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_during_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2a_prov_ls.summary)

res_during_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_during_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_during_step1b_w = fTestWeakInstruments(df_during_fd.dum_ls, res_during_step1_w.fitted_values, res_during_step1b_w.fitted_values, 2) #2.223427716006406

### LS/TA
res_during_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_during_step1b_ls = fTestWeakInstruments(df_during_fd.ls_tot_ta, res_during_step1_ls.fitted_values, res_during_step1b_ls.fitted_values, 2) #0.8311144131052095

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_during_fd['resid_step1_w'] = res_during_step1_w.resids

#----------------------------------------------
#### Charge-off
res_during_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_charge_w_endo.summary) # p-val = 0.0022

#----------------------------------------------
#### Allow
res_during_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_allow_w_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### rwata
res_during_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_rwata_w_endo.summary) # p-val = 0.0025

#----------------------------------------------
#### prov
res_during_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_prov_w_endo.summary) # p-val = 0.9789


#----------------------------------------------
### LS/TA
df_during_fd['resid_step1_ls'] = res_during_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_during_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_charge_ls_endo.summary) # p-val = 0.0321

#----------------------------------------------
#### Allow
res_during_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_allow_ls_endo.summary) # p-val = 0.0006

#----------------------------------------------
#### rwata
res_during_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_rwata_ls_endo.summary) # p-val = 0.0243

#----------------------------------------------
#### prov
res_during_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_during, data = df_during_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_during_prov_ls_endo.summary) # p-val = 0.9036

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_during_step2a_charge_w = sargan(res_during_step2a_charge_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.1689
oir_during_step2b_charge_w = sargan(res_during_step2b_charge_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.2130

#----------------------------------------------
#### Allow
oir_during_step2a_allow_w = sargan(res_during_step2a_allow_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.7000
oir_during_step2b_allow_w = sargan(res_during_step2b_allow_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.8011

#----------------------------------------------
#### rwata
oir_during_step2a_rwata_w = sargan(res_during_step2a_rwata_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.2378
oir_during_step2b_rwata_w = sargan(res_during_step2b_rwata_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.1369

#----------------------------------------------
#### prov
oir_during_step2a_prov_w = sargan(res_during_step2a_prov_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.8507
oir_during_step2b_prov_w = sargan(res_during_step2b_prov_w.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.9606

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_during_step2a_charge_ls = sargan(res_during_step2a_charge_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0049
oir_during_step2b_charge_ls = sargan(res_during_step2b_charge_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0063

#----------------------------------------------
#### Allow
oir_during_step2a_allow_ls = sargan(res_during_step2a_allow_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0058
oir_during_step2b_allow_ls = sargan(res_during_step2b_allow_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0069

#----------------------------------------------
#### rwata
oir_during_step2a_rwata_ls = sargan(res_during_step2a_rwata_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0027
oir_during_step2b_rwata_ls = sargan(res_during_step2b_rwata_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.0030

#----------------------------------------------
#### prov
oir_during_step2a_prov_ls = sargan(res_during_step2a_prov_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.8760
oir_during_step2b_prov_ls = sargan(res_during_step2b_prov_ls.resids, df_during_fd[righthand_x.split(' + ')], df_during_fd[righthand_z.split(' + ')]) # p-val = 0.9156

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 3: Post Crisis
#----------------------------------------------
'''-----------------------------------------''' 

# First check the data on column rank
if log:
    rank_post = np.linalg.matrix_rank(df_post_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_post = np.linalg.matrix_rank(df_post_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
                                
#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat             
## Dummy LS
res_post_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step1_w.summary)
df_post_fd['G_hat_fd_w'] = res_post_step1_w.fitted_values

## LS/TA
res_post_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step1_ls.summary)
df_post_fd['G_hat_fd_ls'] = res_post_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_post_fd.loc[:,df_post_fd.columns.str.contains('_xbar')] * df_post_fd.G_hat_fd_w[:, None]
if log:
    df_post_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_post_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_post_fd.loc[:,df_post_fd.columns.str.contains('_xbar')] * df_post_fd.G_hat_fd_ls[:, None]
if log:
    df_post_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_post_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls

#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_post_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_charge_w.summary)

res_post_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_post_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_allow_w.summary)

res_post_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_post_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_rwata_w.summary)

res_post_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_post_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_prov_w.summary)

res_post_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_post_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_charge_ls.summary)

res_post_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_post_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_allow_ls.summary)

res_post_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_post_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_rwata_ls.summary)

res_post_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_post_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2a_prov_ls.summary)

res_post_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_post_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_post_step1b_w = fTestWeakInstruments(df_post_fd.dum_ls, res_post_step1_w.fitted_values, res_post_step1b_w.fitted_values, 2) #0.7411053849984024

### LS/TA
res_post_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_post_step1b_ls = fTestWeakInstruments(df_post_fd.ls_tot_ta, res_post_step1_ls.fitted_values, res_post_step1b_ls.fitted_values, 2) #12.158629955639562

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_post_fd['resid_step1_w'] = res_post_step1_w.resids

#----------------------------------------------
#### Charge-off
res_post_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_charge_w_endo.summary) # p-val = 0.5143

#----------------------------------------------
#### Allow
res_post_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_allow_w_endo.summary) # p-val = 0.0654

#----------------------------------------------
#### rwata
res_post_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_rwata_w_endo.summary) # p-val = 0.0183

#----------------------------------------------
#### prov
res_post_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_prov_w_endo.summary) # p-val = 0.9713


#----------------------------------------------
### LS/TA
df_post_fd['resid_step1_ls'] = res_post_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_post_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_charge_ls_endo.summary) # p-val = 0.1515

#----------------------------------------------
#### Allow
res_post_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_allow_ls_endo.summary) # p-val = 0.7587

#----------------------------------------------
#### rwata
res_post_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_rwata_ls_endo.summary) # p-val = 0.1396

#----------------------------------------------
#### prov
res_post_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies_post, data = df_post_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_post_prov_ls_endo.summary) # p-val = 0.3057

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_post_step2a_charge_w = sargan(res_post_step2a_charge_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.2873
oir_post_step2b_charge_w = sargan(res_post_step2b_charge_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.2289

#----------------------------------------------
#### Allow
oir_post_step2a_allow_w = sargan(res_post_step2a_allow_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.2772
oir_post_step2b_allow_w = sargan(res_post_step2b_allow_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.1291

#----------------------------------------------
#### rwata
oir_post_step2a_rwata_w = sargan(res_post_step2a_rwata_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.4020
oir_post_step2b_rwata_w = sargan(res_post_step2b_rwata_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.2363

#----------------------------------------------
#### prov
oir_post_step2a_prov_w = sargan(res_post_step2a_prov_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.1869
oir_post_step2b_prov_w = sargan(res_post_step2b_prov_w.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.1852

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_post_step2a_charge_ls = sargan(res_post_step2a_charge_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.7704
oir_post_step2b_charge_ls = sargan(res_post_step2b_charge_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### Allow
oir_post_step2a_allow_ls = sargan(res_post_step2a_allow_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.0108
oir_post_step2b_allow_ls = sargan(res_post_step2b_allow_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### rwata
oir_post_step2a_rwata_ls = sargan(res_post_step2a_rwata_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.0004
oir_post_step2b_rwata_ls = sargan(res_post_step2b_rwata_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.0001

#----------------------------------------------
#### prov
oir_post_step2a_prov_ls = sargan(res_post_step2a_prov_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.4849
oir_post_step2b_prov_ls = sargan(res_post_step2b_prov_ls.resids, df_post_fd[righthand_x.split(' + ')], df_post_fd[righthand_z.split(' + ')]) # p-val = 0.3829

#----------------------------------------------
#----------------------------------------------
# Tables
#----------------------------------------------
#----------------------------------------------
# Prelims
## Make dict that contains all variables and names
dict_var_names = {'distance':'Max Distance Branches',
                 'provratio':'Loan Loss Provisions',
                 'rwata':'RWA/TA',
                 'net_coffratio_tot_ta':'Loan Charge-offs',
                 'allowratio_tot_ta':'Loan Loss Allowances',
                 'ls_tot_ta':'Loan Sales/TA',
                 'dum_ls':'Dummy Loan Sales',
                 'size':'Log(TA)',
                 'RC7205':'Regulatory Capital Ratio',
                 'loanratio':'Loan Ratio',
                 'roa':'ROA',
                 'depratio':'Deposit Ratio',
                 'comloanratio':'Commercial Loan Ratio',
                 'RC2170':'Total Assets',
                 'num_branch':'Num Branches',
                 'bhc':'BHC Indicator',
                 'RIAD4150':'Num Employees',
                 'perc_limited_branch':'Limited Branches (in %)',
                 'perc_full_branch':'Full Branches (in %)',
                 'unique_states':'Num States Active',
                 'UNIT':'Unit Bank Indicator',
                 'G_hat_fd_ls': 'Loan Sales/TA',
                 'G_hat_fd_w':'Dummy Loan Sales'}

### Add the ghat_w variables to the dict
vars_ghat_w = pd.Series(righthand_ghat_w.split(' + ')).unique()
vars_ghat_ls = pd.Series(righthand_ghat_ls.split(' + ')).unique()
vars_x = pd.Series(righthand_x.split(' + ')).unique()

dict_ghat_w = {}
dict_ghat_ls = {}

for key, name in zip(vars_ghat_w, dict_var_names):
    dict_ghat_w[key] = '$\hat{{G}}$({})'.format(dict_var_names[name]) 
for key, name in zip(vars_ghat_ls, dict_var_names):
    dict_ghat_ls[key] = '$\hat{{G}}$({})'.format(dict_var_names[name])
    
dict_var_names.update(dict_ghat_w)
dict_var_names.update(dict_ghat_ls)     

## Regression orders
reg_order_step1 = righthand_z.split(' + ') + ['RC7205','loanratio','roa','depratio','comloanratio','RC2170','bhc']
reg_order_step2_w = ['G_hat_fd_w'] + righthand_x.split(' + ') + righthand_ghat_w.split(' + ')
reg_order_step2_ls = ['G_hat_fd_ls'] + righthand_x.split(' + ') + righthand_ghat_ls.split(' + ')

## Set the var names
var_names_step1 = [dict_var_names[key] for key in reg_order_step1]
var_names_step2_w = [dict_var_names[key] for key in reg_order_step2_w]
var_names_step2_ls = [dict_var_names[key] for key in reg_order_step2_ls]

# Make tables
table_step1_pre = summary_col([res_pre_step1_w,res_pre_step1_ls], show = 'se', regressor_order = reg_order_step1)
table_step1_during = summary_col([res_during_step1_w,res_during_step1_ls], show = 'se', regressor_order = reg_order_step1)
table_step1_post = summary_col([res_post_step1_w,res_post_step1_ls], show = 'se', regressor_order = reg_order_step1)

table_step2_w_pre = summary_col([\
    res_pre_step2a_charge_w, res_pre_step2b_charge_w, \
    res_pre_step2a_allow_w, res_pre_step2b_allow_w,\
    res_pre_step2a_rwata_w, res_pre_step2b_rwata_w, \
    res_pre_step2a_prov_w, res_pre_step2b_prov_w],\
                     show = 'se', regressor_order = var_names_step2_w)
table_step2_w_during = summary_col([\
    res_during_step2a_charge_w, res_during_step2b_charge_w,\
    res_during_step2a_allow_w,res_during_step2b_allow_w,\
    res_during_step2a_rwata_w,res_during_step2b_rwata_w,\
    res_during_step2a_prov_w,res_during_step2b_prov_w],\
                     show = 'se', regressor_order = var_names_step2_w)
table_step2_w_post = summary_col([\
    res_post_step2a_charge_w, res_post_step2b_charge_w,\
    res_post_step2a_allow_w,res_post_step2b_allow_w,\
    res_post_step2a_rwata_w,res_post_step2b_rwata_w,\
    res_post_step2a_prov_w,res_post_step2b_prov_w],\
                     show = 'se', regressor_order = var_names_step2_w)

table_step2_ls_pre = summary_col([\
    res_pre_step2a_charge_ls,res_pre_step2b_charge_ls,\
    res_pre_step2a_allow_ls, res_pre_step2b_allow_ls, \
    res_pre_step2a_rwata_ls, res_pre_step2b_rwata_ls, \
    res_pre_step2a_prov_ls, res_pre_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)
table_step2_ls_during = summary_col([\
    res_during_step2a_charge_ls,res_during_step2b_charge_ls,\
    res_during_step2a_allow_ls,res_during_step2b_allow_ls,\
    res_during_step2a_rwata_ls,res_during_step2b_rwata_ls,\
    res_during_step2a_prov_ls,res_during_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)
table_step2_ls_post = summary_col([\
    res_post_step2a_charge_ls,res_post_step2b_charge_ls,\
    res_post_step2a_allow_ls,res_post_step2b_allow_ls,\
    res_post_step2a_rwata_ls,res_post_step2b_rwata_ls,\
    res_post_step2a_prov_ls,res_post_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)

#----------------------------------------------
#----------------------------------------------
# Statistic Tables
#----------------------------------------------
#----------------------------------------------
# Setup the basic tables
weak_tests_pre = [f_test_pre_step1b_w,f_test_pre_step1b_ls]
weak_tests_during = [f_test_during_step1b_w,f_test_during_step1b_ls]
weak_tests_post = [f_test_post_step1b_w,f_test_post_step1b_ls]

endo_tests_pre = [res_pre_charge_w_endo.pvalues[-1],res_pre_allow_w_endo.pvalues[-1],\
                   res_pre_rwata_w_endo.pvalues[-1],res_pre_prov_w_endo.pvalues[-1],\
                   res_pre_charge_ls_endo.pvalues[-1],res_pre_allow_ls_endo.pvalues[-1],\
                   res_pre_rwata_ls_endo.pvalues[-1],res_pre_prov_ls_endo.pvalues[-1]]
endo_tests_during = [res_during_charge_w_endo.pvalues[-1],res_during_allow_w_endo.pvalues[-1],\
                  res_during_rwata_w_endo.pvalues[-1],res_during_prov_w_endo.pvalues[-1],\
                  res_during_charge_ls_endo.pvalues[-1],res_during_allow_ls_endo.pvalues[-1],\
                  res_during_rwata_ls_endo.pvalues[-1],res_during_prov_ls_endo.pvalues[-1]]
endo_tests_post = [res_post_charge_w_endo.pvalues[-1],res_post_allow_w_endo.pvalues[-1],\
                  res_post_rwata_w_endo.pvalues[-1],res_post_prov_w_endo.pvalues[-1],\
                  res_post_charge_ls_endo.pvalues[-1],res_post_allow_ls_endo.pvalues[-1],\
                  res_post_rwata_ls_endo.pvalues[-1],res_post_prov_ls_endo.pvalues[-1]]

sargan_tests_pre = [oir_pre_step2a_charge_w.pval, oir_pre_step2b_charge_w.pval,\
                              oir_pre_step2a_allow_w.pval, oir_pre_step2b_allow_w.pval,\
                              oir_pre_step2a_rwata_w.pval, oir_pre_step2b_rwata_w.pval,\
                              oir_pre_step2a_prov_w.pval, oir_pre_step2b_prov_w.pval,\
                              oir_pre_step2a_charge_ls.pval, oir_pre_step2b_charge_ls.pval,\
                              oir_pre_step2a_allow_ls.pval, oir_pre_step2b_allow_ls.pval,\
                              oir_pre_step2a_rwata_ls.pval, oir_pre_step2b_rwata_ls.pval,\
                              oir_pre_step2a_prov_ls.pval, oir_pre_step2b_prov_ls.pval]
sargan_tests_during = [oir_during_step2a_charge_w.pval, oir_during_step2b_charge_w.pval,\
                              oir_during_step2a_allow_w.pval, oir_during_step2b_allow_w.pval,\
                              oir_during_step2a_rwata_w.pval, oir_during_step2b_rwata_w.pval,\
                              oir_during_step2a_prov_w.pval, oir_during_step2b_prov_w.pval,\
                              oir_during_step2a_charge_ls.pval, oir_during_step2b_charge_ls.pval,\
                              oir_during_step2a_allow_ls.pval, oir_during_step2b_allow_ls.pval,\
                              oir_during_step2a_rwata_ls.pval, oir_during_step2b_rwata_ls.pval,\
                              oir_during_step2a_prov_ls.pval, oir_during_step2b_prov_ls.pval]
sargan_tests_post = [oir_post_step2a_charge_w.pval, oir_post_step2b_charge_w.pval,\
                              oir_post_step2a_allow_w.pval, oir_post_step2b_allow_w.pval,\
                              oir_post_step2a_rwata_w.pval, oir_post_step2b_rwata_w.pval,\
                              oir_post_step2a_prov_w.pval, oir_post_step2b_prov_w.pval,\
                              oir_post_step2a_charge_ls.pval, oir_post_step2b_charge_ls.pval,\
                              oir_post_step2a_allow_ls.pval, oir_post_step2b_allow_ls.pval,\
                              oir_post_step2a_rwata_ls.pval, oir_post_step2b_rwata_ls.pval,\
                              oir_post_step2a_prov_ls.pval, oir_post_step2b_prov_ls.pval]

# Zip the lists and make one table out of them
weak_tests_pre_zip = [j for i in zip(weak_tests_pre,weak_tests_pre) for j in i]
weak_tests_pre_zip = [j for i in zip([j for i in zip(weak_tests_pre_zip,weak_tests_pre_zip) for j in i],[j for i in zip(weak_tests_pre_zip,weak_tests_pre_zip) for j in i]) for j in i]
weak_tests_during_zip = [j for i in zip(weak_tests_during,weak_tests_during) for j in i]
weak_tests_during_zip = [j for i in zip([j for i in zip(weak_tests_during_zip,weak_tests_during_zip) for j in i],[j for i in zip(weak_tests_during_zip,weak_tests_during_zip) for j in i]) for j in i]
weak_tests_post_zip = [j for i in zip(weak_tests_post,weak_tests_post) for j in i]
weak_tests_post_zip = [j for i in zip([j for i in zip(weak_tests_post_zip,weak_tests_post_zip) for j in i],[j for i in zip(weak_tests_post_zip,weak_tests_post_zip) for j in i]) for j in i]

endo_tests_pre_zip = [j for i in zip(endo_tests_pre,endo_tests_pre) for j in i]
endo_tests_during_zip = [j for i in zip(endo_tests_during,endo_tests_during) for j in i]
endo_tests_post_zip = [j for i in zip(endo_tests_post,endo_tests_post) for j in i]

# Make a row that returns 1 if weak_f > 10, endo < 0.05 and sargan > 0.05
indicator_tests_pre = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_pre_zip], [(i < 0.05) * 1 for i in endo_tests_pre_zip], [(i > 0.05) * 1 for i in sargan_tests_pre])]
indicator_tests_during = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_during_zip], [(i < 0.05) * 1 for i in endo_tests_during_zip], [(i > 0.05) * 1 for i in sargan_tests_during])]
indicator_tests_post = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_post_zip], [(i < 0.05) * 1 for i in endo_tests_post_zip], [(i > 0.05) * 1 for i in sargan_tests_post])]


# Make a pandas dataframe and save to excel
index = ['F-test weak instruments','P-val endogenous w','P-val Sargan','Indicator']
columns_pre = ['charge_2a_pre_w','charge_2b_pre_w','allow_2a_pre_w','allow_2b_pre_w',\
           'rwata_2a_pre_w','rwata_2b_pre_w','prov_2a_pre_w','prov_2b_pre_w',\
           'charge_2a_pre_ls','charge_2b_pre_ls','allow_2a_pre_ls','allow_2b_pre_ls',\
           'rwata_2a_pre_ls','rwata_2b_pre_ls','prov_2a_pre_ls','prov_2b_pre_ls']
columns_during = ['charge_2a_during_w','charge_2b_during_w','allow_2a_during_w','allow_2b_during_w',\
           'rwata_2a_during_w','rwata_2b_during_w','prov_2a_during_w','prov_2b_during_w',\
           'charge_2a_during_ls','charge_2b_during_ls','allow_2a_during_ls','allow_2b_during_ls',\
           'rwata_2a_during_ls','rwata_2b_during_ls','prov_2a_during_ls','prov_2b_during_ls']
columns_post = ['charge_2a_post_w','charge_2b_post_w','allow_2a_post_w','allow_2b_post_w',\
           'rwata_2a_post_w','rwata_2b_post_w','prov_2a_post_w','prov_2b_post_w',\
           'charge_2a_post_ls','charge_2b_post_ls','allow_2a_post_ls','allow_2b_post_ls',\
           'rwata_2a_post_ls','rwata_2b_post_ls','prov_2a_post_ls','prov_2b_post_ls']


df_tests_pre = pd.DataFrame([weak_tests_pre_zip,endo_tests_pre_zip,sargan_tests_pre,indicator_tests_pre], index = index, columns = columns_pre)
df_tests_during = pd.DataFrame([weak_tests_during_zip,endo_tests_during_zip,sargan_tests_during,indicator_tests_during], index = index, columns = columns_during)
df_tests_post = pd.DataFrame([weak_tests_post_zip,endo_tests_post_zip,sargan_tests_post,indicator_tests_post], index = index, columns = columns_post)

#-----------------------------------------------
# Save to a single excel
from openpyxl import load_workbook
path = r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char\FD_IV_results.xlsx'

rename_index_step1 = dict(zip(reg_order_step1,var_names_step1))
rename_index_step2_w = dict(zip(reg_order_step2_w,var_names_step2_w))
rename_index_step2_ls = dict(zip(reg_order_step2_ls,var_names_step2_ls))

if log:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_pre.to_excel(writer, sheet_name = 'precrisis_step1_log', rename_index = rename_index_step1)
        table_step2_w_pre.to_excel(writer, sheet_name = 'precrisis_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_pre.to_excel(writer, sheet_name = 'precrisis_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_pre.to_excel(writer, sheet_name = 'precrisis_tests_log')
        table_step1_during.to_excel(writer, sheet_name = 'crisis_step1_log', rename_index = rename_index_step1)
        table_step2_w_during.to_excel(writer, sheet_name = 'crisis_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_during.to_excel(writer, sheet_name = 'crisis_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_during.to_excel(writer, sheet_name = 'crisis_tests_log')
        table_step1_post.to_excel(writer, sheet_name = 'postcrisis_step1_log', rename_index = rename_index_step1)
        table_step2_w_post.to_excel(writer, sheet_name = 'postcrisis_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_post.to_excel(writer, sheet_name = 'postcrisis_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_post.to_excel(writer, sheet_name = 'postcrisis_tests_log')
        
        writer.save()
        writer.close()  
else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_pre.to_excel(writer, sheet_name = 'precrisis_step1', rename_index = rename_index_step1)
        table_step2_w_pre.to_excel(writer, sheet_name = 'precrisis_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_pre.to_excel(writer, sheet_name = 'precrisis_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_pre.to_excel(writer, sheet_name = 'precrisis_tests')
        table_step1_during.to_excel(writer, sheet_name = 'crisis_step1', rename_index = rename_index_step1)
        table_step2_w_during.to_excel(writer, sheet_name = 'crisis_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_during.to_excel(writer, sheet_name = 'crisis_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_during.to_excel(writer, sheet_name = 'crisis_tests')
        table_step1_post.to_excel(writer, sheet_name = 'postcrisis_step1', rename_index = rename_index_step1)
        table_step2_w_post.to_excel(writer, sheet_name = 'postcrisis_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_post.to_excel(writer, sheet_name = 'postcrisis_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_post.to_excel(writer, sheet_name = 'postcrisis_tests')
        
        writer.save()
        writer.close()  