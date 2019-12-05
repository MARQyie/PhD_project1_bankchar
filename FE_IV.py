#------------------------------------------
# IV treatment model for first working paper
# Mark van der Plaat
# October 2019 

 # Import packages
import pandas as pd
import numpy as np
import scipy.stats as sps # used to calculated cdf and pdf

import os
os.chdir(r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char')

# Import method that adds a constant to a df
from statsmodels.tools.tools import add_constant

# Import method that can estimate a pooled probit
from statsmodels.discrete.discrete_model import Probit

# Import method for POLS (also does FE)
from linearmodels import PanelOLS

# Import packages for the Sargan-Hausman test
from linearmodels.iv._utility import annihilate
from linearmodels.utility import WaldTestStatistic

import sys # to use the help functions needed
sys.path.insert(1, r'X:\My Documents\PhD\Coding_docs\Help_functions')

from summary3 import summary_col
#--------------------------------------------
# Set parameters 
log = True # If set to False the program estimates the model without logs and with size
change_ls = True # If set to False the program will run a different subset and append it to the excel

#--------------------------------------------
''' This script estimates the treatment effect of loan sales on credit risk
    with an IV estimation procedure. The procedure has two steps
    
    Step 1: Estimate a probit I(LS > 0) on 1, X and Z, where X are the exogenous
    variables and Z are the instruments. Obtain the fitted probabilities G()
    
    Step 2: Do a OLS of CR on 1, G_hat, X, G_hat(X-X_bar)
    
    The first model does not explicitely correct for fixed effects.
    
    Dynamic effects are not included.
    '''   
    
#---------------------------------------------- 
#----------------------------------------------
# Prelims
#----------------------------------------------
#----------------------------------------------

#----------------------------------------------
# Load data and add needed variables

# Load df
df = pd.read_csv('df_wp1_clean.csv', index_col = 0)

## Make multi index
df.date = pd.to_datetime(df.date.astype(str).str.strip() + '1230')
df.set_index(['IDRSSD','date'],inplace=True)

## Drop missings on distance
df.dropna(subset = ['distance'], inplace = True)

## Dummy variable for loan sales
if log:
    df['dum_ls'] = np.exp((df.ls_tot > 0) * 1) - 1 #will be taken the log of later
else:
    df['dum_ls'] = (df.ls_tot > 0) * 1    

## Take a subset of variables (only the ones needed)
vars_needed = ['distance','provratio','rwata','net_coffratio_tot_ta',\
               'allowratio_tot_ta','ls_tot_ta','dum_ls','size',\
               'RC7205','loanratio','roa','depratio','comloanratio','RC2170',\
               'num_branch', 'bhc', 'RIAD4150', 'perc_limited_branch',\
               'perc_full_branch', 'unique_states','UNIT']
df_full = df[vars_needed]

## drop NaNs
df_full.dropna(subset = ['provratio','rwata','net_coffratio_tot_ta','allowratio_tot_ta',\
               'ls_tot_ta','RC7205','loanratio','roa',\
               'depratio','comloanratio','RC2170','size'], inplace = True)

#---------------------------------------------------
# Setup the data

## Set aside TA
ta = df_full.RC2170

## Correct dummy and percentage variables for log
if log:
    df_full['bhc'] = np.exp(df_full.bhc) - 1
    df_full['UNIT'] = np.exp(df_full.UNIT) - 1

## Take logs of the df
if log:
    df_full = df_full.transform(lambda df: np.log(1 + df)).replace([np.inf, -np.inf], 0)

## Add TA for subsetting
df_full['ta_sub'] = ta

## Add the x_xbar to the df
if log:
    x_xbar = df_full[['RC7205','loanratio','roa',\
                      'depratio','comloanratio','RC2170','bhc']].transform(lambda df: df - df.mean())
    df_full[[x + '_xbar' for x in ['RC7205','loanratio',\
                                   'roa','depratio','comloanratio','RC2170','bhc']]] = x_xbar
else:
    x_xbar = df_full[['RC7205','loanratio','roa',\
                          'depratio','comloanratio','size','bhc']].transform(lambda df: df - df.mean())
    df_full[[x + '_xbar' for x in ['RC7205','loanratio',\
                                   'roa','depratio','comloanratio','size','bhc']]] = x_xbar

# Subset the df 
## Only take the banks that change in dum_ls
if change_ls:
    intersect = np.intersect1d(df_full[df_full.ls_tot_ta > 0].index.\
                               get_level_values(0).unique(),\
                               df_full[df_full.ls_tot_ta == 0].index.\
                               get_level_values(0).unique())
    df_sub = df_full[df_full.index.get_level_values(0).isin(intersect)]
else:
    ## Kick out the community banks (based on Stiroh, 2004)  
    ids_comm = df_full[((df_full.index.get_level_values(1) == pd.Timestamp(2018,12,30)) &\
                     (df_full.ta_sub < 3e5) & (df_full.bhc == 0))].index.get_level_values(0).unique().tolist() 
    ids_tot = df_full.index.get_level_values(0).unique().tolist()
    ids_sub = [x for x in ids_tot if x not in ids_comm]
    df_sub = df_full[df_full.index.get_level_values(0).isin(ids_sub)]   

## Add dummies
dummy_full_fe = pd.get_dummies(df_full.index.get_level_values(1))
dummy_sub_fe = pd.get_dummies(df_sub.index.get_level_values(1))

### Add dummies to the dfs
col_dummy = ['dum' + dummy for dummy in dummy_full_fe.columns.astype(str).str[:4].tolist()]
dummy_full = pd.DataFrame(np.array(dummy_full_fe), index = df_full.index, columns = col_dummy)
dummy_sub = pd.DataFrame(np.array(dummy_sub_fe), index = df_sub.index, columns = col_dummy)
df_full[col_dummy] = dummy_full
df_sub[col_dummy] = dummy_sub

## Transform to fixed effects
df_full_fe = df_full.groupby(df_full.index.get_level_values(0)).transform(lambda df: df - df.mean()).dropna()
df_sub_fe = df_sub.groupby(df_sub.index.get_level_values(0)).transform(lambda df: df - df.mean()).dropna()

#---------------------------------------------------
# Load the necessary functions

def fTestWeakInstruments(y, fitted_full, fitted_reduced, dof = 4):
    ''' Simple F-test to test the strength of instrumental variables.
        
        y : True y values
        fitted_full : fitted values of the first stage with instruments
        fitted_reduced : fitted values first stage without instruments
        dof : number of instruments'''
    
    # Calculate the SSE and MSE
    sse_full = np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    sse_reduced =  np.sum([(y.values[i] - fitted_reduced.values[i][0])**2 for i in range(y.shape[0])])
    
    mse_full = (1 / y.shape[0]) * np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    
    # Calculate the statistic
    f_stat = ((sse_reduced - sse_full)/dof) / mse_full
    
    return f_stat

def sargan(resids, x, z, nendog = 1):
    '''Function performs a sargan test (no heteroskedasity) to check
        the validity of the overidentification restrictions. H0: 
        overidentifying restrictions hold.
        
        resids : residuals of the second stage
        x : exogenous variables
        z : instruments 
        nendog : number of endogenous variables'''  

    nobs, ninstr = z.shape
    name = 'Sargan\'s test of overidentification'

    eps = resids.values[:,None]
    u = annihilate(eps, pd.concat([x,z],axis = 1))
    stat = nobs * (1 - (u.T @ u) / (eps.T @ eps)).squeeze()
    null = 'The overidentification restrictions are valid'

    return WaldTestStatistic(stat, null, ninstr - nendog, name=name)

#----------------------------------------------
#----------------------------------------------
# Analyses
#----------------------------------------------
#----------------------------------------------

# Set the righthand side of the formulas
if log:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + RC2170 + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + RC2170_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + RC2170_G_hat_ls + bhc_G_hat_ls'
else:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + size + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + size_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + size_G_hat_ls + bhc_G_hat_ls' 
    
righthand_z = r'unique_states + RIAD4150'
num_z = righthand_z.count('+') + 1 # Use string count as a trick to count the number of vars in z
time_dummies = ' + '.join(col_dummy[1:])

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 1: Full Sample
#----------------------------------------------
'''-----------------------------------------''' 

# First check the data on column rank
if log:
    rank_full = np.linalg.matrix_rank(df_full_fe[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_full = np.linalg.matrix_rank(df_full_fe[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'size','num_branch','bhc']]) # 10 (should be 10)

#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat
                
## Dummy LS
res_full_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step1_w.summary)
df_full_fe['G_hat_fe_w'] = res_full_step1_w.fitted_values

## LS/TA
res_full_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step1_ls.summary)
df_full_fe['G_hat_fe_ls'] = res_full_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fe_w = df_full_fe.loc[:,df_full_fe.columns.str.contains('_xbar')] * df_full_fe.G_hat_fe_w[:, None]
if log:
    df_full_fe[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fe_w
else:
    df_full_fe[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fe_w

G_hat_x_xbar_fe_ls = df_full_fe.loc[:,df_full_fe.columns.str.contains('_xbar')] * df_full_fe.G_hat_fe_ls[:, None]
if log:
    df_full_fe[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fe_ls
else:
    df_full_fe[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fe_ls
    
#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_full_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_charge_w.summary)

res_full_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_full_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_allow_w.summary)

res_full_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_full_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_rwata_w.summary)

res_full_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_full_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_prov_w.summary)

res_full_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_full_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_charge_ls.summary)

res_full_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_full_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_allow_ls.summary)

res_full_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_full_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_rwata_ls.summary)

res_full_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_full_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_prov_ls.summary)

res_full_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_full_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
f_test_full_step1b_w = fTestWeakInstruments(df_full_fe.dum_ls, res_full_step1_w.fitted_values, res_full_step1b_w.fitted_values, 2) #11.141455804993024

### LS/TA
res_full_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
f_test_full_step1b_ls = fTestWeakInstruments(df_full_fe.ls_tot_ta, res_full_step1_ls.fitted_values, res_full_step1b_ls.fitted_values, 2) #17.200081157277296

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_full_fe['resid_step1_w'] = res_full_step1_w.resids

#----------------------------------------------
#### Charge-off
res_full_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_charge_w_endo.summary) # p-val = 0.0407

#----------------------------------------------
#### Allow
res_full_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_allow_w_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### rwata
res_full_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_rwata_w_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### prov
res_full_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_prov_w_endo.summary) # p-val = 0.0681

#----------------------------------------------
### LS/TA
df_full_fe['resid_step1_ls'] = res_full_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_full_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_charge_ls_endo.summary) # p-val = 0.0296

#----------------------------------------------
#### Allow
res_full_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_allow_ls_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### rwata
res_full_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_rwata_ls_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### prov
res_full_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_prov_ls_endo.summary) # p-val = 0.0493

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_full_step2a_charge_w = sargan(res_full_step2a_charge_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0000
oir_full_step2b_charge_w = sargan(res_full_step2b_charge_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### Allow
oir_full_step2a_allow_w = sargan(res_full_step2a_allow_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0000
oir_full_step2b_allow_w = sargan(res_full_step2b_allow_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### rwata
oir_full_step2a_rwata_w = sargan(res_full_step2a_rwata_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0023
oir_full_step2b_rwata_w = sargan(res_full_step2b_rwata_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0015

#----------------------------------------------
#### prov
oir_full_step2a_prov_w = sargan(res_full_step2a_prov_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.6458
oir_full_step2b_prov_w = sargan(res_full_step2b_prov_w.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.5889

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_full_step2a_charge_ls = sargan(res_full_step2a_charge_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0049
oir_full_step2b_charge_ls = sargan(res_full_step2b_charge_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0063

#----------------------------------------------
#### Allow
oir_full_step2a_allow_ls = sargan(res_full_step2a_allow_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0058
oir_full_step2b_allow_ls = sargan(res_full_step2b_allow_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0069

#----------------------------------------------
#### rwata
oir_full_step2a_rwata_ls = sargan(res_full_step2a_rwata_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0027
oir_full_step2b_rwata_ls = sargan(res_full_step2b_rwata_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.0030

#----------------------------------------------
#### prov
oir_full_step2a_prov_ls = sargan(res_full_step2a_prov_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.8760
oir_full_step2b_prov_ls = sargan(res_full_step2b_prov_ls.resids, df_full_fe[righthand_x.split(' + ')], df_full_fe[righthand_z.split(' + ')]) # p-val = 0.9156

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 2: Subsample
#----------------------------------------------
'''-----------------------------------------''' 

# First check the data on column rank
if log:
    rank_sub = np.linalg.matrix_rank(df_sub_fe[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_sub = np.linalg.matrix_rank(df_sub_fe[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'size','num_branch','bhc']]) # 10 (should be 10)

#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat             
## Dummy LS
res_sub_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step1_w.summary)
df_sub_fe['G_hat_fe_w'] = res_sub_step1_w.fitted_values

## LS/TA
res_sub_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step1_ls.summary)
df_sub_fe['G_hat_fe_ls'] = res_sub_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fe_w = df_sub_fe.loc[:,df_sub_fe.columns.str.contains('_xbar')] * df_sub_fe.G_hat_fe_w[:, None]
if log:
    df_sub_fe[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                    'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fe_w
else:
    df_sub_fe[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                    'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fe_w

G_hat_x_xbar_fe_ls = df_sub_fe.loc[:,df_sub_fe.columns.str.contains('_xbar')] * df_sub_fe.G_hat_fe_ls[:, None]
if log:
    df_sub_fe[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                    'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fe_ls
else:
    df_sub_fe[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                    'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fe_ls
#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_sub_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_charge_w.summary)

res_sub_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_sub_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_allow_w.summary)

res_sub_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_sub_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_rwata_w.summary)

res_sub_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_sub_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_prov_w.summary)

res_sub_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_sub_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_charge_ls.summary)

res_sub_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_sub_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_allow_ls.summary)

res_sub_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_sub_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_rwata_ls.summary)

res_sub_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_sub_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_prov_ls.summary)

res_sub_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fe_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_sub_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
f_test_sub_step1b_w = fTestWeakInstruments(df_sub_fe.dum_ls, res_sub_step1_w.fitted_values, res_sub_step1b_w.fitted_values, 2) #0.7411053849984024

### LS/TA
res_sub_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
f_test_sub_step1b_ls = fTestWeakInstruments(df_sub_fe.ls_tot_ta, res_sub_step1_ls.fitted_values, res_sub_step1b_ls.fitted_values, 2) #12.158629955639562

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_sub_fe['resid_step1_w'] = res_sub_step1_w.resids

#----------------------------------------------
#### Charge-off
res_sub_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_charge_w_endo.summary) # p-val = 0.5143

#----------------------------------------------
#### Allow
res_sub_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_allow_w_endo.summary) # p-val = 0.0654

#----------------------------------------------
#### rwata
res_sub_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_rwata_w_endo.summary) # p-val = 0.0183

#----------------------------------------------
#### prov
res_sub_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_prov_w_endo.summary) # p-val = 0.9713


#----------------------------------------------
### LS/TA
df_sub_fe['resid_step1_ls'] = res_sub_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_sub_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_charge_ls_endo.summary) # p-val = 0.1515

#----------------------------------------------
#### Allow
res_sub_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_allow_ls_endo.summary) # p-val = 0.7587

#----------------------------------------------
#### rwata
res_sub_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_rwata_ls_endo.summary) # p-val = 0.1396

#----------------------------------------------
#### prov
res_sub_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_prov_ls_endo.summary) # p-val = 0.3057

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_sub_step2a_charge_w = sargan(res_sub_step2a_charge_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.2873
oir_sub_step2b_charge_w = sargan(res_sub_step2b_charge_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.2289

#----------------------------------------------
#### Allow
oir_sub_step2a_allow_w = sargan(res_sub_step2a_allow_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.2772
oir_sub_step2b_allow_w = sargan(res_sub_step2b_allow_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.1291

#----------------------------------------------
#### rwata
oir_sub_step2a_rwata_w = sargan(res_sub_step2a_rwata_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.4020
oir_sub_step2b_rwata_w = sargan(res_sub_step2b_rwata_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.2363

#----------------------------------------------
#### prov
oir_sub_step2a_prov_w = sargan(res_sub_step2a_prov_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.1869
oir_sub_step2b_prov_w = sargan(res_sub_step2b_prov_w.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.1852

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_sub_step2a_charge_ls = sargan(res_sub_step2a_charge_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.7704
oir_sub_step2b_charge_ls = sargan(res_sub_step2b_charge_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### Allow
oir_sub_step2a_allow_ls = sargan(res_sub_step2a_allow_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.0108
oir_sub_step2b_allow_ls = sargan(res_sub_step2b_allow_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### rwata
oir_sub_step2a_rwata_ls = sargan(res_sub_step2a_rwata_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.0004
oir_sub_step2b_rwata_ls = sargan(res_sub_step2b_rwata_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.0001

#----------------------------------------------
#### prov
oir_sub_step2a_prov_ls = sargan(res_sub_step2a_prov_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.4849
oir_sub_step2b_prov_ls = sargan(res_sub_step2b_prov_ls.resids, df_sub_fe[righthand_x.split(' + ')], df_sub_fe[righthand_z.split(' + ')]) # p-val = 0.3829

#----------------------------------------------
#----------------------------------------------
# Regression Tables
#----------------------------------------------
#----------------------------------------------\
# Prelims
## Make dict that contains all variables and names
dict_var_names = {'distance':'Max Distance Branches',
                 'provratio':'Loan Loss Provisions',
                 'rwata':'RWA/TA',
                 'net_coffratio_tot_ta':'Loan Charge-offs',
                 'allowratio_tot_ta':'Loan Loss Allowances',
                 'ls_tot_ta':'Loan Sales/TA',
                 'dum_ls':'Dummy Loan Sales',
                 'size':'Log(TA)',
                 'RC7205':'Regulatory Capital Ratio',
                 'loanratio':'Loan Ratio',
                 'roa':'ROA',
                 'depratio':'Deposit Ratio',
                 'comloanratio':'Commercial Loan Ratio',
                 'RC2170':'Total Assets',
                 'num_branch':'Num Branches',
                 'bhc':'BHC Indicator',
                 'RIAD4150':'Num Employees',
                 'perc_limited_branch':'Limited Branches (in %)',
                 'perc_full_branch':'Full Branches (in %)',
                 'unique_states':'Num States Active',
                 'UNIT':'Unit Bank Indicator',
                 'G_hat_fe_ls': 'Loan Sales/TA',
                 'G_hat_fe_w':'Dummy Loan Sales'}

### Add the ghat_w variables to the dict
vars_ghat_w = pd.Series(righthand_ghat_w.split(' + ')).unique()
vars_ghat_ls = pd.Series(righthand_ghat_ls.split(' + ')).unique()
vars_x = pd.Series(righthand_x.split(' + ')).unique()

dict_ghat_w = {}
dict_ghat_ls = {}

for key, name in zip(vars_ghat_w, dict_var_names):
    dict_ghat_w[key] = '$\hat{{G}}$({})'.format(dict_var_names[name]) 
for key, name in zip(vars_ghat_ls, dict_var_names):
    dict_ghat_ls[key] = '$\hat{{G}}$({})'.format(dict_var_names[name])
    
dict_var_names.update(dict_ghat_w)
dict_var_names.update(dict_ghat_ls)     

## Regression orders
reg_order_step1 = righthand_z.split(' + ') + ['RC7205','loanratio','roa','depratio','comloanratio','RC2170','bhc']
reg_order_step2_w = ['G_hat_fe_w'] + righthand_x.split(' + ') + righthand_ghat_w.split(' + ')
reg_order_step2_ls = ['G_hat_fe_ls'] + righthand_x.split(' + ') + righthand_ghat_ls.split(' + ')

## Set the var names
var_names_step1 = [dict_var_names[key] for key in reg_order_step1]
var_names_step2_w = [dict_var_names[key] for key in reg_order_step2_w]
var_names_step2_ls = [dict_var_names[key] for key in reg_order_step2_ls]
# Make tables
table_step1_full = summary_col([res_full_step1_w,res_full_step1_ls], show = 'se', regressor_order = reg_order_step1)
table_step1_sub = summary_col([res_sub_step1_w,res_sub_step1_ls], show = 'se', regressor_order = reg_order_step1)

table_step2_w_full = summary_col([\
    res_full_step2a_charge_w, res_full_step2b_charge_w, \
    res_full_step2a_allow_w, res_full_step2b_allow_w,\
    res_full_step2a_rwata_w, res_full_step2b_rwata_w, \
    res_full_step2a_prov_w, res_full_step2b_prov_w],\
                     show = 'se', regressor_order = reg_order_step2_w)
table_step2_w_sub = summary_col([\
    res_sub_step2a_charge_w, res_sub_step2b_charge_w,\
    res_sub_step2a_allow_w,res_sub_step2b_allow_w,\
    res_sub_step2a_rwata_w,res_sub_step2b_rwata_w,\
    res_sub_step2a_prov_w,res_sub_step2b_prov_w],\
                     show = 'se', regressor_order = reg_order_step2_w)

table_step2_ls_full = summary_col([\
    res_full_step2a_charge_ls,res_full_step2b_charge_ls,\
    res_full_step2a_allow_ls, res_full_step2b_allow_ls, \
    res_full_step2a_rwata_ls, res_full_step2b_rwata_ls, \
    res_full_step2a_prov_ls, res_full_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)
table_step2_ls_sub = summary_col([\
    res_sub_step2a_charge_ls,res_sub_step2b_charge_ls,\
    res_sub_step2a_allow_ls,res_sub_step2b_allow_ls,\
    res_sub_step2a_rwata_ls,res_sub_step2b_rwata_ls,\
    res_sub_step2a_prov_ls,res_sub_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)

#----------------------------------------------
#----------------------------------------------
# Statistic Tables
#----------------------------------------------
#----------------------------------------------
# Setup the basic tables
weak_tests_full = [f_test_full_step1b_w,f_test_full_step1b_ls]
weak_tests_sub = [f_test_sub_step1b_w,f_test_sub_step1b_ls]

endo_tests_full = [res_full_charge_w_endo.pvalues[-1],res_full_allow_w_endo.pvalues[-1],\
                   res_full_rwata_w_endo.pvalues[-1],res_full_prov_w_endo.pvalues[-1],\
                   res_full_charge_ls_endo.pvalues[-1],res_full_allow_ls_endo.pvalues[-1],\
                   res_full_rwata_ls_endo.pvalues[-1],res_full_prov_ls_endo.pvalues[-1]]
endo_tests_sub = [res_sub_charge_w_endo.pvalues[-1],res_sub_allow_w_endo.pvalues[-1],\
                  res_sub_rwata_w_endo.pvalues[-1],res_sub_prov_w_endo.pvalues[-1],\
                  res_sub_charge_ls_endo.pvalues[-1],res_sub_allow_ls_endo.pvalues[-1],\
                  res_sub_rwata_ls_endo.pvalues[-1],res_sub_prov_ls_endo.pvalues[-1]]


sargan_tests_full = [oir_full_step2a_charge_w.pval, oir_full_step2b_charge_w.pval,\
                              oir_full_step2a_allow_w.pval, oir_full_step2b_allow_w.pval,\
                              oir_full_step2a_rwata_w.pval, oir_full_step2b_rwata_w.pval,\
                              oir_full_step2a_prov_w.pval, oir_full_step2b_prov_w.pval,\
                              oir_full_step2a_charge_ls.pval, oir_full_step2b_charge_ls.pval,\
                              oir_full_step2a_allow_ls.pval, oir_full_step2b_allow_ls.pval,\
                              oir_full_step2a_rwata_ls.pval, oir_full_step2b_rwata_ls.pval,\
                              oir_full_step2a_prov_ls.pval, oir_full_step2b_prov_ls.pval]
sargan_tests_sub = [oir_sub_step2a_charge_w.pval, oir_sub_step2b_charge_w.pval,\
                              oir_sub_step2a_allow_w.pval, oir_sub_step2b_allow_w.pval,\
                              oir_sub_step2a_rwata_w.pval, oir_sub_step2b_rwata_w.pval,\
                              oir_sub_step2a_prov_w.pval, oir_sub_step2b_prov_w.pval,\
                              oir_sub_step2a_charge_ls.pval, oir_sub_step2b_charge_ls.pval,\
                              oir_sub_step2a_allow_ls.pval, oir_sub_step2b_allow_ls.pval,\
                              oir_sub_step2a_rwata_ls.pval, oir_sub_step2b_rwata_ls.pval,\
                              oir_sub_step2a_prov_ls.pval, oir_sub_step2b_prov_ls.pval]

# Zip the lists and make one table out of them
weak_tests_full_zip = [j for i in zip(weak_tests_full,weak_tests_full) for j in i]
weak_tests_full_zip = [j for i in zip([j for i in zip(weak_tests_full_zip,weak_tests_full_zip) for j in i],[j for i in zip(weak_tests_full_zip,weak_tests_full_zip) for j in i]) for j in i]
weak_tests_sub_zip = [j for i in zip(weak_tests_sub,weak_tests_sub) for j in i]
weak_tests_sub_zip = [j for i in zip([j for i in zip(weak_tests_sub_zip,weak_tests_sub_zip) for j in i],[j for i in zip(weak_tests_sub_zip,weak_tests_sub_zip) for j in i]) for j in i]

endo_tests_full_zip = [j for i in zip(endo_tests_full,endo_tests_full) for j in i]
endo_tests_sub_zip = [j for i in zip(endo_tests_sub,endo_tests_sub) for j in i]

# Make a row that returns 1 if weak_f > 10, endo < 0.05 and sargan > 0.05
indicator_tests_full = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_full_zip], [(i < 0.05) * 1 for i in endo_tests_full_zip], [(i > 0.05) * 1 for i in sargan_tests_full])]
indicator_tests_sub = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_sub_zip], [(i < 0.05) * 1 for i in endo_tests_sub_zip], [(i > 0.05) * 1 for i in sargan_tests_sub])]


# Make a pandas dataframe and save to excel
index = ['F-test weak instruments','P-val endogenous w','P-val Sargan','Indicator']
columns_full = ['charge_2a_full_w','charge_2b_full_w','allow_2a_full_w','allow_2b_full_w',\
           'rwata_2a_full_w','rwata_2b_full_w','prov_2a_full_w','prov_2b_full_w',\
           'charge_2a_full_ls','charge_2b_full_ls','allow_2a_full_ls','allow_2b_full_ls',\
           'rwata_2a_full_ls','rwata_2b_full_ls','prov_2a_full_ls','prov_2b_full_ls']
columns_sub = ['charge_2a_sub_w','charge_2b_sub_w','allow_2a_sub_w','allow_2b_sub_w',\
           'rwata_2a_sub_w','rwata_2b_sub_w','prov_2a_sub_w','prov_2b_sub_w',\
           'charge_2a_sub_ls','charge_2b_sub_ls','allow_2a_sub_ls','allow_2b_sub_ls',\
           'rwata_2a_sub_ls','rwata_2b_sub_ls','prov_2a_sub_ls','prov_2b_sub_ls']


df_tests_full = pd.DataFrame([weak_tests_full_zip,endo_tests_full_zip,sargan_tests_full,indicator_tests_full], index = index, columns = columns_full)
df_tests_sub = pd.DataFrame([weak_tests_sub_zip,endo_tests_sub_zip,sargan_tests_sub,indicator_tests_sub], index = index, columns = columns_sub)

#-----------------------------------------------
# Save to a single excel
from openpyxl import load_workbook
path = r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char\FE_IV_results.xlsx'

rename_index_step1 = dict(zip(reg_order_step1,var_names_step1))
rename_index_step2_w = dict(zip(reg_order_step2_w,var_names_step2_w))
rename_index_step2_ls = dict(zip(reg_order_step2_ls,var_names_step2_ls))

if log:
    if change_ls:
        with pd.ExcelWriter('FE_IV_results.xlsx') as writer:
            table_step1_full.to_excel(writer, sheet_name = 'Full_step1_log', rename_index = rename_index_step1)
            table_step2_w_full.to_excel(writer, sheet_name = 'Full_step2_w_log', rename_index = rename_index_step2_w)
            table_step2_ls_full.to_excel(writer, sheet_name = 'Full_step2_ls_log', rename_index = rename_index_step2_ls)
            df_tests_full.to_excel(writer, sheet_name = 'Full_tests_log')
            table_step1_sub.to_excel(writer, sheet_name = 'Sub_changels_step1_log', rename_index = rename_index_step1)
            table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_w_log', rename_index = rename_index_step2_w)
            table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_ls_log', rename_index = rename_index_step2_ls)
            df_tests_sub.to_excel(writer, sheet_name = 'Sub_changels_tests_log')
    else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step1_log', rename_index = rename_index_step1)
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_sub.to_excel(writer, sheet_name = 'Sub_nocomm_tests_log')
        
        writer.save()
        writer.close()  
           
else:
    if change_ls:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        
        table_step1_full.to_excel(writer, sheet_name = 'Full_step1', rename_index = rename_index_step1)
        table_step2_w_full.to_excel(writer, sheet_name = 'Full_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_full.to_excel(writer, sheet_name = 'Full_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_full.to_excel(writer, sheet_name = 'Full_tests_log')
        table_step1_sub.to_excel(writer, sheet_name = 'Sub_changels_step1', rename_index = rename_index_step1)
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_sub.to_excel(writer, sheet_name = 'Sub_changels_tests')
    
        writer.save()
        writer.close()
    else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step1', rename_index = rename_index_step1)
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_sub.to_excel(writer, sheet_name = 'Sub_nocomm_tests')
        
        writer.save()
        writer.close()  
                       