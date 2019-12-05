#------------------------------------------
# IV models based on different TA size for first working paper
# Mark van der Plaat
# October 2019 

 # Import packages
import pandas as pd
import numpy as np
import scipy.stats as sps # used to calculated cdf and pdf

import os
os.chdir(r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char')

# Import method that adds a constant to a df
from statsmodels.tools.tools import add_constant

# Import method that can estimate a pooled probit
from statsmodels.discrete.discrete_model import Probit

# Import method for POLS (also does FE)
from linearmodels import PanelOLS

# Import packages for the Sargan-Hausman test
from linearmodels.iv._utility import annihilate
from linearmodels.utility import WaldTestStatistic

import sys # to use the help functions needed
sys.path.insert(1, r'X:\My Documents\PhD\Coding_docs\Help_functions')

from summary3 import summary_col

#--------------------------------------------
# Set parameters 
log = False # If set to False the program estimates the model without logs and with size

#--------------------------------------------
''' This script estimates the treatment effect of loan sales on credit risk
    with an IV estimation procedure. The procedure has two steps
    
    Step 1: Estimate a probit I(LS > 0) on 1, X and Z, where X are the exogenous
    variables and Z are the instruments. Obtain the fitted probabilities G()
    
    Step 2: Do a OLS of CR on 1, G_hat, X, G_hat(X-X_bar)
    
    The first model does not explicitely correct for fixed effects.
    
    Dynamic effects are not included.
    '''  
#---------------------------------------------- 
#----------------------------------------------
# Prelims
#----------------------------------------------
#----------------------------------------------

#----------------------------------------------
# Load data and add needed variables

# Load df
df = pd.read_csv('df_wp1_clean.csv', index_col = 0)

## Make multi index
df.date = pd.to_datetime(df.date.astype(str).str.strip() + '1230')
df.set_index(['IDRSSD','date'],inplace=True)

## Drop missings on distance
df.dropna(subset = ['distance'], inplace = True)

## Dummy variable for loan sales
if log:
    df['dum_ls'] = np.exp((df.ls_tot > 0) * 1) - 1 #will be taken the log of later
else:
    df['dum_ls'] = (df.ls_tot > 0) * 1  

## Take a subset of variables (only the ones needed)
vars_needed = ['distance','provratio','rwata','net_coffratio_tot_ta',\
               'allowratio_tot_ta','ls_tot_ta','dum_ls','size',\
               'RC7205','loanratio','roa','depratio','comloanratio','RC2170',\
               'num_branch', 'bhc', 'RIAD4150', 'perc_limited_branch',\
               'perc_full_branch', 'unique_states','UNIT']
df_sub = df[vars_needed]

## drop NaNs
df_sub.dropna(subset = ['provratio','rwata','net_coffratio_tot_ta','allowratio_tot_ta',\
               'ls_tot_ta','RC7205','loanratio','roa',\
               'depratio','comloanratio','RC2170','size'], inplace = True)
#---------------------------------------------------
# Setup the data

## Set aside TA
ta = df_sub.RC2170

## Take logs of the df
if log:
    df_sub = df_sub.transform(lambda df: np.log(1 + df)).replace([np.inf, -np.inf], 0)

## Add TA for subsetting
df_sub['ta_sub'] = ta

## Add the x_xbar to the df
if log:
    x_xbar = df_sub[['RC7205','loanratio','roa',\
                      'depratio','comloanratio','RC2170','bhc']].transform(lambda df: df - df.mean())
    df_sub[[x + '_xbar' for x in ['RC7205','loanratio','roa',\
                      'depratio','comloanratio','RC2170','bhc']]] = x_xbar
else:
    x_xbar = df_sub[['RC7205','loanratio','roa',\
                          'depratio','comloanratio','size','bhc']].transform(lambda df: df - df.mean())
    df_sub[[x + '_xbar' for x in ['RC7205','loanratio','roa',\
                          'depratio','comloanratio','size','bhc']]] = x_xbar
# Subset the df
'''We take the Total assets in 2018 to split the sample into three.
    Small banks have TA < 300,000,000
    Medium banks have TA in (300,000,000; 1,000,000,000)
    Large banks have TA > 1,000,000,000'''
    
ids_small = df_sub[(df_sub.index.get_level_values(1) == pd.Timestamp(2018,12,30)) & (df_sub.ta_sub < 3e5)].index.get_level_values(0).unique().tolist()
ids_medium = df_sub[(df_sub.index.get_level_values(1) == pd.Timestamp(2018,12,30)) & (df_sub.ta_sub.between(3e5,1e6))].index.get_level_values(0).unique().tolist()
ids_large = df_sub[(df_sub.index.get_level_values(1) == pd.Timestamp(2018,12,30)) & (df_sub.ta_sub > 1e6)].index.get_level_values(0).unique().tolist()

df_small = df_sub[df_sub.index.get_level_values(0).isin(ids_small)]
df_med = df_sub[df_sub.index.get_level_values(0).isin(ids_medium)]
df_large = df_sub[df_sub.index.get_level_values(0).isin(ids_large)]

### Check the sets on doubles
'''NOTE: All intersects are zero, continue'''
intersec_smallmed = np.intersect1d(ids_small,ids_medium)
intersec_smalllarge = np.intersect1d(ids_small,ids_large)
intersec_medlarge = np.intersect1d(ids_medium,ids_large)

## Take the first differences
df_small_fd = df_small.groupby(df_small.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()
df_med_fd = df_med.groupby(df_med.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()
df_large_fd = df_large.groupby(df_large.index.get_level_values(0)).transform(lambda df: df.shift(periods = 1) - df).dropna()

## Add dummies
dummy_small_fd = pd.get_dummies(df_small_fd.index.get_level_values(1))
dummy_med_fd = pd.get_dummies(df_med_fd.index.get_level_values(1))
dummy_large_fd = pd.get_dummies(df_large_fd.index.get_level_values(1))

### Add dummies to the dfs
col_dummy = ['dum' + dummy for dummy in dummy_small_fd.columns.astype(str).str[:4].tolist()]
dummy_small_fd = pd.DataFrame(np.array(dummy_small_fd), index = df_small_fd.index, columns = col_dummy)
dummy_med_fd = pd.DataFrame(np.array(dummy_med_fd), index = df_med_fd.index, columns = col_dummy)
dummy_large_fd = pd.DataFrame(np.array(dummy_large_fd), index = df_large_fd.index, columns = col_dummy)

df_small_fd[col_dummy] = dummy_small_fd
df_med_fd[col_dummy] = dummy_med_fd
df_large_fd[col_dummy] = dummy_large_fd

#---------------------------------------------------
# Load the necessary functions

def fTestWeakInstruments(y, fitted_full, fitted_reduced, dof = 4):
    ''' Simple F-test to test the strength of instrumental variables.
        
        y : True y values
        fitted_full : fitted values of the first stage with instruments
        fitted_reduced : fitted values first stage without instruments
        dof : number of instruments'''
    
    # Calculate the SSE and MSE
    sse_full = np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    sse_reduced =  np.sum([(y.values[i] - fitted_reduced.values[i][0])**2 for i in range(y.shape[0])])
    
    mse_full = (1 / y.shape[0]) * np.sum([(y.values[i] - fitted_full.values[i][0])**2 for i in range(y.shape[0])])
    
    # Calculate the statistic
    f_stat = ((sse_reduced - sse_full)/dof) / mse_full
    
    return f_stat

def sargan(resids, x, z, nendog = 1):
    '''Function performs a sargan test (no heteroskedasity) to check
        the validity of the overidentification restrictions. H0: 
        overidentifying restrictions hold.
        
        resids : residuals of the second stage
        x : exogenous variables
        z : instruments 
        nendog : number of endogenous variables'''  

    nobs, ninstr = z.shape
    name = 'Sargan\'s test of overidentification'

    eps = resids.values[:,None]
    u = annihilate(eps, pd.concat([x,z],axis = 1))
    stat = nobs * (1 - (u.T @ u) / (eps.T @ eps)).squeeze()
    null = 'The overidentification restrictions are valid'

    return WaldTestStatistic(stat, null, ninstr - nendog, name=name)

#----------------------------------------------
#----------------------------------------------
# Analyses
#----------------------------------------------
#----------------------------------------------

if log:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + RC2170 + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + RC2170_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + RC2170_G_hat_ls + bhc_G_hat_ls'
else:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + size + bhc'
    righthand_ghat_w = r'RC7205_G_hat_w + loanratio_G_hat_w + roa_G_hat_w + depratio_G_hat_w + comloanratio_G_hat_w + size_G_hat_w + bhc_G_hat_w'
    righthand_ghat_ls = r'RC7205_G_hat_ls + loanratio_G_hat_ls + roa_G_hat_ls + depratio_G_hat_ls + comloanratio_G_hat_ls + size_G_hat_ls + bhc_G_hat_ls' 
    
righthand_z = r'unique_states + RIAD4150'
num_z = righthand_z.count('+') + 1 # Use string count as a trick to count the number of vars in z
time_dummies = ' + '.join(col_dummy[1:])

#----------------------------------------------
# MODEL 1: SMALL BANKS, ls_dum
#----------------------------------------------

# First check the data on column rank
if log:
    rank_small = np.linalg.matrix_rank(df_small_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_small = np.linalg.matrix_rank(df_small_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'size','num_branch','bhc']]) # 10 (should be 10)

#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat
                
## Dummy LS
res_small_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step1_w.summary)
df_small_fd['G_hat_fd_w'] = res_small_step1_w.fitted_values

## LS/TA
res_small_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step1_ls.summary)
df_small_fd['G_hat_fd_ls'] = res_small_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_small_fd.loc[:,df_small_fd.columns.str.contains('_xbar')] * df_small_fd.G_hat_fd_w[:, None]
if log:
    df_small_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_small_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_small_fd.loc[:,df_small_fd.columns.str.contains('_xbar')] * df_small_fd.G_hat_fd_ls[:, None]
if log:
    df_small_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_small_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls
    
#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_small_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_charge_w.summary)

res_small_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_small_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_allow_w.summary)

res_small_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_small_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_rwata_w.summary)

res_small_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_small_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_prov_w.summary)

res_small_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_small_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_charge_ls.summary)

res_small_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_small_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_allow_ls.summary)

res_small_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_small_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_rwata_ls.summary)

res_small_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_small_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2a_prov_ls.summary)

res_small_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_small_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_small_step1b_w = fTestWeakInstruments(df_small_fd.dum_ls, res_small_step1_w.fitted_values, res_small_step1b_w.fitted_values, 2) 

### LS/TA
res_small_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_small_step1b_ls = fTestWeakInstruments(df_small_fd.ls_tot_ta, res_small_step1_ls.fitted_values, res_small_step1b_ls.fitted_values, 2)

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_small_fd['resid_step1_w'] = res_small_step1_w.resids

#----------------------------------------------
#### Charge-off
res_small_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_charge_w_endo.summary) 

#----------------------------------------------
#### Allow
res_small_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_allow_w_endo.summary) 

#----------------------------------------------
#### rwata
res_small_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_rwata_w_endo.summary)

#----------------------------------------------
#### prov
res_small_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_prov_w_endo.summary) 

#----------------------------------------------
### LS/TA
df_small_fd['resid_step1_ls'] = res_small_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_small_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_charge_ls_endo.summary)

#----------------------------------------------
#### Allow
res_small_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_allow_ls_endo.summary) 

#----------------------------------------------
#### rwata
res_small_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_rwata_ls_endo.summary) 

#----------------------------------------------
#### prov
res_small_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_small_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_small_prov_ls_endo.summary) 

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_small_step2a_charge_w = sargan(res_small_step2a_charge_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0000
oir_small_step2b_charge_w = sargan(res_small_step2b_charge_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### Allow
oir_small_step2a_allow_w = sargan(res_small_step2a_allow_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0000
oir_small_step2b_allow_w = sargan(res_small_step2b_allow_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0000

#----------------------------------------------
#### rwata
oir_small_step2a_rwata_w = sargan(res_small_step2a_rwata_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0023
oir_small_step2b_rwata_w = sargan(res_small_step2b_rwata_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0015

#----------------------------------------------
#### prov
oir_small_step2a_prov_w = sargan(res_small_step2a_prov_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.6458
oir_small_step2b_prov_w = sargan(res_small_step2b_prov_w.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.5889

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_small_step2a_charge_ls = sargan(res_small_step2a_charge_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0049
oir_small_step2b_charge_ls = sargan(res_small_step2b_charge_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0063

#----------------------------------------------
#### Allow
oir_small_step2a_allow_ls = sargan(res_small_step2a_allow_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0058
oir_small_step2b_allow_ls = sargan(res_small_step2b_allow_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0069

#----------------------------------------------
#### rwata
oir_small_step2a_rwata_ls = sargan(res_small_step2a_rwata_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0027
oir_small_step2b_rwata_ls = sargan(res_small_step2b_rwata_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.0030

#----------------------------------------------
#### prov
oir_small_step2a_prov_ls = sargan(res_small_step2a_prov_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.8760
oir_small_step2b_prov_ls = sargan(res_small_step2b_prov_ls.resids, df_small_fd[righthand_x.split(' + ')], df_small_fd[righthand_z.split(' + ')]) # p-val = 0.9156


'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 2: Medium Banks
#----------------------------------------------
'''-----------------------------------------''' 

if log:
    rank_med = np.linalg.matrix_rank(df_med_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_med = np.linalg.matrix_rank(df_med_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
          
#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat
                
## Dummy LS
res_med_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step1_w.summary)
df_med_fd['G_hat_fd_w'] = res_med_step1_w.fitted_values

## LS/TA
res_med_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step1_ls.summary)
df_med_fd['G_hat_fd_ls'] = res_med_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_med_fd.loc[:,df_med_fd.columns.str.contains('_xbar')] * df_med_fd.G_hat_fd_w[:, None]
if log:
    df_med_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_med_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_med_fd.loc[:,df_med_fd.columns.str.contains('_xbar')] * df_med_fd.G_hat_fd_ls[:, None]
if log:
    df_med_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_med_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls

#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_med_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_charge_w.summary)

res_med_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_med_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_allow_w.summary)

res_med_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_med_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_rwata_w.summary)

res_med_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_med_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_prov_w.summary)

res_med_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_med_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_charge_ls.summary)

res_med_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_med_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_allow_ls.summary)

res_med_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_med_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_rwata_ls.summary)

res_med_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_med_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2a_prov_ls.summary)

res_med_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_med_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_med_step1b_w = fTestWeakInstruments(df_med_fd.dum_ls, res_med_step1_w.fitted_values, res_med_step1b_w.fitted_values, 2) #2.223427716006406

### LS/TA
res_med_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_med_step1b_ls = fTestWeakInstruments(df_med_fd.ls_tot_ta, res_med_step1_ls.fitted_values, res_med_step1b_ls.fitted_values, 2) #0.8311144131052095

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_med_fd['resid_step1_w'] = res_med_step1_w.resids

#----------------------------------------------
#### Charge-off
res_med_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_charge_w_endo.summary) # p-val = 0.0022

#----------------------------------------------
#### Allow
res_med_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_allow_w_endo.summary) # p-val = 0.0000

#----------------------------------------------
#### rwata
res_med_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_rwata_w_endo.summary) # p-val = 0.0025

#----------------------------------------------
#### prov
res_med_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_prov_w_endo.summary) # p-val = 0.9789


#----------------------------------------------
### LS/TA
df_med_fd['resid_step1_ls'] = res_med_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_med_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_charge_ls_endo.summary) # p-val = 0.0321

#----------------------------------------------
#### Allow
res_med_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_allow_ls_endo.summary) # p-val = 0.0006

#----------------------------------------------
#### rwata
res_med_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_rwata_ls_endo.summary) # p-val = 0.0243

#----------------------------------------------
#### prov
res_med_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_med_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_med_prov_ls_endo.summary) # p-val = 0.9036

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_med_step2a_charge_w = sargan(res_med_step2a_charge_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.1689
oir_med_step2b_charge_w = sargan(res_med_step2b_charge_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.2130

#----------------------------------------------
#### Allow
oir_med_step2a_allow_w = sargan(res_med_step2a_allow_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.7000
oir_med_step2b_allow_w = sargan(res_med_step2b_allow_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.8011

#----------------------------------------------
#### rwata
oir_med_step2a_rwata_w = sargan(res_med_step2a_rwata_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.2378
oir_med_step2b_rwata_w = sargan(res_med_step2b_rwata_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.1369

#----------------------------------------------
#### prov
oir_med_step2a_prov_w = sargan(res_med_step2a_prov_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.8507
oir_med_step2b_prov_w = sargan(res_med_step2b_prov_w.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.9606

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_med_step2a_charge_ls = sargan(res_med_step2a_charge_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0049
oir_med_step2b_charge_ls = sargan(res_med_step2b_charge_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0063

#----------------------------------------------
#### Allow
oir_med_step2a_allow_ls = sargan(res_med_step2a_allow_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0058
oir_med_step2b_allow_ls = sargan(res_med_step2b_allow_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0069

#----------------------------------------------
#### rwata
oir_med_step2a_rwata_ls = sargan(res_med_step2a_rwata_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0027
oir_med_step2b_rwata_ls = sargan(res_med_step2b_rwata_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.0030

#----------------------------------------------
#### prov
oir_med_step2a_prov_ls = sargan(res_med_step2a_prov_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.8760
oir_med_step2b_prov_ls = sargan(res_med_step2b_prov_ls.resids, df_med_fd[righthand_x.split(' + ')], df_med_fd[righthand_z.split(' + ')]) # p-val = 0.9156

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 3: large Banks
#----------------------------------------------
'''-----------------------------------------''' 

# First check the data on column rank
if log:
    rank_large = np.linalg.matrix_rank(df_large_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
else:
    rank_large = np.linalg.matrix_rank(df_large_fd[[\
                              'RC7205','loanratio','roa','depratio','comloanratio',\
                              'RC2170','num_branch','bhc']]) # 10 (should be 10)
                                
#----------------------------------------------
# STEP 1: First Stage
#----------------------------------------------
                            
# Estimate G_hat             
## Dummy LS
res_large_step1_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step1_w.summary)
df_large_fd['G_hat_fd_w'] = res_large_step1_w.fitted_values

## LS/TA
res_large_step1_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + righthand_z + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step1_ls.summary)
df_large_fd['G_hat_fd_ls'] = res_large_step1_ls.fitted_values

#----------------------------------------------
# Calculate G_hat_x_xbar for both first stages
G_hat_x_xbar_fd_w = df_large_fd.loc[:,df_large_fd.columns.str.contains('_xbar')] * df_large_fd.G_hat_fd_w[:, None]
if log:
    df_large_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_w
else:
    df_large_fd[[x + '_G_hat_w' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_w

G_hat_x_xbar_fd_ls = df_large_fd.loc[:,df_large_fd.columns.str.contains('_xbar')] * df_large_fd.G_hat_fd_ls[:, None]
if log:
    df_large_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','RC2170','bhc']]] = G_hat_x_xbar_fd_ls
else:
    df_large_fd[[x + '_G_hat_ls' for x in ['RC7205','loanratio','roa',\
                                         'depratio','comloanratio','size','bhc']]] = G_hat_x_xbar_fd_ls

#----------------------------------------------
# Step 2: Second Stage
#----------------------------------------------

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_large_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_charge_w.summary)

res_large_step2b_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_charge_w.summary)

#----------------------------------------------
## Allowance 
res_large_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_allow_w.summary)

res_large_step2b_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_allow_w.summary)

#----------------------------------------------
## rwata 
res_large_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_rwata_w.summary)

res_large_step2b_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_rwata_w.summary)

#----------------------------------------------
## prov 
res_large_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_prov_w.summary)

res_large_step2b_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_w' + '+' + righthand_ghat_w + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_large_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_charge_ls.summary)

res_large_step2b_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_charge_ls.summary)

#----------------------------------------------
## Allowance
res_large_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_allow_ls.summary)

res_large_step2b_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_allow_ls.summary)

#----------------------------------------------
## rwata
res_large_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_rwata_ls.summary)

res_large_step2b_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_rwata_ls.summary)

#----------------------------------------------
## prov
res_large_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2a_prov_ls.summary)

res_large_step2b_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'G_hat_fd_ls' + '+' + righthand_ghat_ls + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_step2b_prov_ls.summary)

#----------------------------------------------
# Tests
'''We test for three things:
    1) The strength of the instrument using a DWH test. F-stat must be > 10
    2) A test whether dum_ls or ls_tot_ta is endogenous. H0: variable is exogenous
    3) A Sargan test to test the overidentifying restrictions. H0: overidentifying restrictions hold'''
#----------------------------------------------

#----------------------------------------------                                
## Weak instruments
### Dummy LS
res_large_step1b_w = PanelOLS.from_formula('dum_ls' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_large_step1b_w = fTestWeakInstruments(df_large_fd.dum_ls, res_large_step1_w.fitted_values, res_large_step1b_w.fitted_values, 2) #0.7411053849984024

### LS/TA
res_large_step1b_ls = PanelOLS.from_formula('ls_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
f_test_large_step1b_ls = fTestWeakInstruments(df_large_fd.ls_tot_ta, res_large_step1_ls.fitted_values, res_large_step1b_ls.fitted_values, 2) #12.158629955639562

#----------------------------------------------
## Endogenous loan sales variable

#----------------------------------------------
### Dummy LS
df_large_fd['resid_step1_w'] = res_large_step1_w.resids

#----------------------------------------------
#### Charge-off
res_large_charge_w_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_charge_w_endo.summary) # p-val = 0.5143

#----------------------------------------------
#### Allow
res_large_allow_w_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_allow_w_endo.summary) # p-val = 0.0654

#----------------------------------------------
#### rwata
res_large_rwata_w_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_rwata_w_endo.summary) # p-val = 0.0183

#----------------------------------------------
#### prov
res_large_prov_w_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'dum_ls' + ' + ' + 'resid_step1_w' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_prov_w_endo.summary) # p-val = 0.9713


#----------------------------------------------
### LS/TA
df_large_fd['resid_step1_ls'] = res_large_step1_ls.resids

#----------------------------------------------
#### Charge-off
res_large_charge_ls_endo = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_charge_ls_endo.summary) # p-val = 0.1515

#----------------------------------------------
#### Allow
res_large_allow_ls_endo = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_allow_ls_endo.summary) # p-val = 0.7587

#----------------------------------------------
#### rwata
res_large_rwata_ls_endo = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_rwata_ls_endo.summary) # p-val = 0.1396

#----------------------------------------------
#### prov
res_large_prov_ls_endo = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + 'ls_tot_ta' + ' + ' + 'resid_step1_ls' + ' + ' + time_dummies, data = df_large_fd).fit(cov_type = 'clustered', cluster_entity = True)
print(res_large_prov_ls_endo.summary) # p-val = 0.3057

#----------------------------------------------
## Sargan test

#----------------------------------------------
### Dummy LS

#----------------------------------------------
#### Charge-off
oir_large_step2a_charge_w = sargan(res_large_step2a_charge_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.2873
oir_large_step2b_charge_w = sargan(res_large_step2b_charge_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.2289

#----------------------------------------------
#### Allow
oir_large_step2a_allow_w = sargan(res_large_step2a_allow_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.2772
oir_large_step2b_allow_w = sargan(res_large_step2b_allow_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.1291

#----------------------------------------------
#### rwata
oir_large_step2a_rwata_w = sargan(res_large_step2a_rwata_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.4020
oir_large_step2b_rwata_w = sargan(res_large_step2b_rwata_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.2363

#----------------------------------------------
#### prov
oir_large_step2a_prov_w = sargan(res_large_step2a_prov_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.1869
oir_large_step2b_prov_w = sargan(res_large_step2b_prov_w.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.1852

#----------------------------------------------
###LS/TA

#----------------------------------------------
#### Charge-off
oir_large_step2a_charge_ls = sargan(res_large_step2a_charge_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.7704
oir_large_step2b_charge_ls = sargan(res_large_step2b_charge_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### Allow
oir_large_step2a_allow_ls = sargan(res_large_step2a_allow_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.0108
oir_large_step2b_allow_ls = sargan(res_large_step2b_allow_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.0313

#----------------------------------------------
#### rwata
oir_large_step2a_rwata_ls = sargan(res_large_step2a_rwata_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.0004
oir_large_step2b_rwata_ls = sargan(res_large_step2b_rwata_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.0001

#----------------------------------------------
#### prov
oir_large_step2a_prov_ls = sargan(res_large_step2a_prov_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.4849
oir_large_step2b_prov_ls = sargan(res_large_step2b_prov_ls.resids, df_large_fd[righthand_x.split(' + ')], df_large_fd[righthand_z.split(' + ')]) # p-val = 0.3829

#----------------------------------------------
#----------------------------------------------
# Tables
#----------------------------------------------
#----------------------------------------------
# Prelims
## Make dict that contains all variables and names
dict_var_names = {'distance':'Max Distance Branches',
                 'provratio':'Loan Loss Provisions',
                 'rwata':'RWA/TA',
                 'net_coffratio_tot_ta':'Loan Charge-offs',
                 'allowratio_tot_ta':'Loan Loss Allowances',
                 'ls_tot_ta':'Loan Sales/TA',
                 'dum_ls':'Dummy Loan Sales',
                 'size':'Log(TA)',
                 'RC7205':'Regulatory Capital Ratio',
                 'loanratio':'Loan Ratio',
                 'roa':'ROA',
                 'depratio':'Deposit Ratio',
                 'comloanratio':'Commercial Loan Ratio',
                 'RC2170':'Total Assets',
                 'num_branch':'Num Branches',
                 'bhc':'BHC Indicator',
                 'RIAD4150':'Num Employees',
                 'perc_limited_branch':'Limited Branches (in %)',
                 'perc_full_branch':'Full Branches (in %)',
                 'unique_states':'Num States Active',
                 'UNIT':'Unit Bank Indicator',
                 'G_hat_fd_ls': 'Loan Sales/TA',
                 'G_hat_fd_w':'Dummy Loan Sales'}

### Add the ghat_w variables to the dict
vars_ghat_w = pd.Series(righthand_ghat_w.split(' + ')).unique()
vars_ghat_ls = pd.Series(righthand_ghat_ls.split(' + ')).unique()
vars_x = pd.Series(righthand_x.split(' + ')).unique()

dict_ghat_w = {}
dict_ghat_ls = {}

for key, name in zip(vars_ghat_w, dict_var_names):
    dict_ghat_w[key] = '$\hat{{G}}$({})'.format(dict_var_names[name]) 
for key, name in zip(vars_ghat_ls, dict_var_names):
    dict_ghat_ls[key] = '$\hat{{G}}$({})'.format(dict_var_names[name])
    
dict_var_names.update(dict_ghat_w)
dict_var_names.update(dict_ghat_ls)     

## Regression orders
reg_order_step1 = righthand_z.split(' + ') + ['RC7205','loanratio','roa','depratio','comloanratio','RC2170','bhc']
reg_order_step2_w = ['G_hat_fd_w'] + righthand_x.split(' + ') + righthand_ghat_w.split(' + ')
reg_order_step2_ls = ['G_hat_fd_ls'] + righthand_x.split(' + ') + righthand_ghat_ls.split(' + ')

## Set the var names
var_names_step1 = [dict_var_names[key] for key in reg_order_step1]
var_names_step2_w = [dict_var_names[key] for key in reg_order_step2_w]
var_names_step2_ls = [dict_var_names[key] for key in reg_order_step2_ls]

# Make tables
table_step1_small = summary_col([res_small_step1_w,res_small_step1_ls], show = 'se', regressor_order = reg_order_step1)
table_step1_med = summary_col([res_med_step1_w,res_med_step1_ls], show = 'se', regressor_order = reg_order_step1)
table_step1_large = summary_col([res_large_step1_w,res_large_step1_ls], show = 'se', regressor_order = reg_order_step1)

table_step2_w_small = summary_col([\
    res_small_step2a_charge_w, res_small_step2b_charge_w, \
    res_small_step2a_allow_w, res_small_step2b_allow_w,\
    res_small_step2a_rwata_w, res_small_step2b_rwata_w, \
    res_small_step2a_prov_w, res_small_step2b_prov_w],\
                     show = 'se', regressor_order = reg_order_step2_w)
table_step2_w_med = summary_col([\
    res_med_step2a_charge_w, res_med_step2b_charge_w,\
    res_med_step2a_allow_w,res_med_step2b_allow_w,\
    res_med_step2a_rwata_w,res_med_step2b_rwata_w,\
    res_med_step2a_prov_w,res_med_step2b_prov_w],\
                     show = 'se', regressor_order = reg_order_step2_w)
table_step2_w_large = summary_col([\
    res_large_step2a_charge_w, res_large_step2b_charge_w,\
    res_large_step2a_allow_w,res_large_step2b_allow_w,\
    res_large_step2a_rwata_w,res_large_step2b_rwata_w,\
    res_large_step2a_prov_w,res_large_step2b_prov_w],\
                     show = 'se', regressor_order = reg_order_step2_w)

table_step2_ls_small = summary_col([\
    res_small_step2a_charge_ls,res_small_step2b_charge_ls,\
    res_small_step2a_allow_ls, res_small_step2b_allow_ls, \
    res_small_step2a_rwata_ls, res_small_step2b_rwata_ls, \
    res_small_step2a_prov_ls, res_small_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)
table_step2_ls_med = summary_col([\
    res_med_step2a_charge_ls,res_med_step2b_charge_ls,\
    res_med_step2a_allow_ls,res_med_step2b_allow_ls,\
    res_med_step2a_rwata_ls,res_med_step2b_rwata_ls,\
    res_med_step2a_prov_ls,res_med_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)
table_step2_ls_large = summary_col([\
    res_large_step2a_charge_ls,res_large_step2b_charge_ls,\
    res_large_step2a_allow_ls,res_large_step2b_allow_ls,\
    res_large_step2a_rwata_ls,res_large_step2b_rwata_ls,\
    res_large_step2a_prov_ls,res_large_step2b_prov_ls],\
                     show = 'se', regressor_order = reg_order_step2_ls)

#----------------------------------------------
#----------------------------------------------
# Statistic Tables
#----------------------------------------------
#----------------------------------------------
# Setup the basic tables
weak_tests_small = [f_test_small_step1b_w,f_test_small_step1b_ls]
weak_tests_med = [f_test_med_step1b_w,f_test_med_step1b_ls]
weak_tests_large = [f_test_large_step1b_w,f_test_large_step1b_ls]

endo_tests_small = [res_small_charge_w_endo.pvalues[-1],res_small_allow_w_endo.pvalues[-1],\
                   res_small_rwata_w_endo.pvalues[-1],res_small_prov_w_endo.pvalues[-1],\
                   res_small_charge_ls_endo.pvalues[-1],res_small_allow_ls_endo.pvalues[-1],\
                   res_small_rwata_ls_endo.pvalues[-1],res_small_prov_ls_endo.pvalues[-1]]
endo_tests_med = [res_med_charge_w_endo.pvalues[-1],res_med_allow_w_endo.pvalues[-1],\
                  res_med_rwata_w_endo.pvalues[-1],res_med_prov_w_endo.pvalues[-1],\
                  res_med_charge_ls_endo.pvalues[-1],res_med_allow_ls_endo.pvalues[-1],\
                  res_med_rwata_ls_endo.pvalues[-1],res_med_prov_ls_endo.pvalues[-1]]
endo_tests_large = [res_large_charge_w_endo.pvalues[-1],res_large_allow_w_endo.pvalues[-1],\
                  res_large_rwata_w_endo.pvalues[-1],res_large_prov_w_endo.pvalues[-1],\
                  res_large_charge_ls_endo.pvalues[-1],res_large_allow_ls_endo.pvalues[-1],\
                  res_large_rwata_ls_endo.pvalues[-1],res_large_prov_ls_endo.pvalues[-1]]

sargan_tests_small = [oir_small_step2a_charge_w.pval, oir_small_step2b_charge_w.pval,\
                              oir_small_step2a_allow_w.pval, oir_small_step2b_allow_w.pval,\
                              oir_small_step2a_rwata_w.pval, oir_small_step2b_rwata_w.pval,\
                              oir_small_step2a_prov_w.pval, oir_small_step2b_prov_w.pval,\
                              oir_small_step2a_charge_ls.pval, oir_small_step2b_charge_ls.pval,\
                              oir_small_step2a_allow_ls.pval, oir_small_step2b_allow_ls.pval,\
                              oir_small_step2a_rwata_ls.pval, oir_small_step2b_rwata_ls.pval,\
                              oir_small_step2a_prov_ls.pval, oir_small_step2b_prov_ls.pval]
sargan_tests_med = [oir_med_step2a_charge_w.pval, oir_med_step2b_charge_w.pval,\
                              oir_med_step2a_allow_w.pval, oir_med_step2b_allow_w.pval,\
                              oir_med_step2a_rwata_w.pval, oir_med_step2b_rwata_w.pval,\
                              oir_med_step2a_prov_w.pval, oir_med_step2b_prov_w.pval,\
                              oir_med_step2a_charge_ls.pval, oir_med_step2b_charge_ls.pval,\
                              oir_med_step2a_allow_ls.pval, oir_med_step2b_allow_ls.pval,\
                              oir_med_step2a_rwata_ls.pval, oir_med_step2b_rwata_ls.pval,\
                              oir_med_step2a_prov_ls.pval, oir_med_step2b_prov_ls.pval]
sargan_tests_large = [oir_large_step2a_charge_w.pval, oir_large_step2b_charge_w.pval,\
                              oir_large_step2a_allow_w.pval, oir_large_step2b_allow_w.pval,\
                              oir_large_step2a_rwata_w.pval, oir_large_step2b_rwata_w.pval,\
                              oir_large_step2a_prov_w.pval, oir_large_step2b_prov_w.pval,\
                              oir_large_step2a_charge_ls.pval, oir_large_step2b_charge_ls.pval,\
                              oir_large_step2a_allow_ls.pval, oir_large_step2b_allow_ls.pval,\
                              oir_large_step2a_rwata_ls.pval, oir_large_step2b_rwata_ls.pval,\
                              oir_large_step2a_prov_ls.pval, oir_large_step2b_prov_ls.pval]

# Zip the lists and make one table out of them
weak_tests_small_zip = [j for i in zip(weak_tests_small,weak_tests_small) for j in i]
weak_tests_small_zip = [j for i in zip([j for i in zip(weak_tests_small_zip,weak_tests_small_zip) for j in i],[j for i in zip(weak_tests_small_zip,weak_tests_small_zip) for j in i]) for j in i]
weak_tests_med_zip = [j for i in zip(weak_tests_med,weak_tests_med) for j in i]
weak_tests_med_zip = [j for i in zip([j for i in zip(weak_tests_med_zip,weak_tests_med_zip) for j in i],[j for i in zip(weak_tests_med_zip,weak_tests_med_zip) for j in i]) for j in i]
weak_tests_large_zip = [j for i in zip(weak_tests_large,weak_tests_large) for j in i]
weak_tests_large_zip = [j for i in zip([j for i in zip(weak_tests_large_zip,weak_tests_large_zip) for j in i],[j for i in zip(weak_tests_large_zip,weak_tests_large_zip) for j in i]) for j in i]

endo_tests_small_zip = [j for i in zip(endo_tests_small,endo_tests_small) for j in i]
endo_tests_med_zip = [j for i in zip(endo_tests_med,endo_tests_med) for j in i]
endo_tests_large_zip = [j for i in zip(endo_tests_large,endo_tests_large) for j in i]

# Make a row that returns 1 if weak_f > 10, endo < 0.05 and sargan > 0.05
indicator_tests_small = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_small_zip], [(i < 0.05) * 1 for i in endo_tests_small_zip], [(i > 0.05) * 1 for i in sargan_tests_small])]
indicator_tests_med = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_med_zip], [(i < 0.05) * 1 for i in endo_tests_med_zip], [(i > 0.05) * 1 for i in sargan_tests_med])]
indicator_tests_large = [a*b*c for a,b,c in zip([(i > 10) * 1 for i in weak_tests_large_zip], [(i < 0.05) * 1 for i in endo_tests_large_zip], [(i > 0.05) * 1 for i in sargan_tests_large])]


# Make a pandas dataframe and save to excel
index = ['F-test weak instruments','P-val endogenous w','P-val Sargan','Indicator']
columns_small = ['charge_2a_small_w','charge_2b_small_w','allow_2a_small_w','allow_2b_small_w',\
           'rwata_2a_small_w','rwata_2b_small_w','prov_2a_small_w','prov_2b_small_w',\
           'charge_2a_small_ls','charge_2b_small_ls','allow_2a_small_ls','allow_2b_small_ls',\
           'rwata_2a_small_ls','rwata_2b_small_ls','prov_2a_small_ls','prov_2b_small_ls']
columns_med = ['charge_2a_med_w','charge_2b_med_w','allow_2a_med_w','allow_2b_med_w',\
           'rwata_2a_med_w','rwata_2b_med_w','prov_2a_med_w','prov_2b_med_w',\
           'charge_2a_med_ls','charge_2b_med_ls','allow_2a_med_ls','allow_2b_med_ls',\
           'rwata_2a_med_ls','rwata_2b_med_ls','prov_2a_med_ls','prov_2b_med_ls']
columns_large = ['charge_2a_large_w','charge_2b_large_w','allow_2a_large_w','allow_2b_large_w',\
           'rwata_2a_large_w','rwata_2b_large_w','prov_2a_large_w','prov_2b_large_w',\
           'charge_2a_large_ls','charge_2b_large_ls','allow_2a_large_ls','allow_2b_large_ls',\
           'rwata_2a_large_ls','rwata_2b_large_ls','prov_2a_large_ls','prov_2b_large_ls']


df_tests_small = pd.DataFrame([weak_tests_small_zip,endo_tests_small_zip,sargan_tests_small,indicator_tests_small], index = index, columns = columns_small)
df_tests_med = pd.DataFrame([weak_tests_med_zip,endo_tests_med_zip,sargan_tests_med,indicator_tests_med], index = index, columns = columns_med)
df_tests_large = pd.DataFrame([weak_tests_large_zip,endo_tests_large_zip,sargan_tests_large,indicator_tests_large], index = index, columns = columns_large)

#-----------------------------------------------
# Save to a single excel
from openpyxl import load_workbook
path = r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char\FD_IV_results.xlsx'

rename_index_step1 = dict(zip(reg_order_step1,var_names_step1))
rename_index_step2_w = dict(zip(reg_order_step2_w,var_names_step2_w))
rename_index_step2_ls = dict(zip(reg_order_step2_ls,var_names_step2_ls))

if log:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_small.to_excel(writer, sheet_name = 'small_step1_log', rename_index = rename_index_step1)
        table_step2_w_small.to_excel(writer, sheet_name = 'small_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_small.to_excel(writer, sheet_name = 'small_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_small.to_excel(writer, sheet_name = 'small_tests_log')
        table_step1_med.to_excel(writer, sheet_name = 'med_step1_log', rename_index = rename_index_step1)
        table_step2_w_med.to_excel(writer, sheet_name = 'med_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_med.to_excel(writer, sheet_name = 'med_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_med.to_excel(writer, sheet_name = 'med_tests_log')
        table_step1_large.to_excel(writer, sheet_name = 'large_step1_log', rename_index = rename_index_step1)
        table_step2_w_large.to_excel(writer, sheet_name = 'large_step2_w_log', rename_index = rename_index_step2_w)
        table_step2_ls_large.to_excel(writer, sheet_name = 'large_step2_ls_log', rename_index = rename_index_step2_ls)
        df_tests_large.to_excel(writer, sheet_name = 'large_tests_log')
        
        writer.save()
        writer.close()  
else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step1_small.to_excel(writer, sheet_name = 'small_step1', rename_index = rename_index_step1)
        table_step2_w_small.to_excel(writer, sheet_name = 'small_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_small.to_excel(writer, sheet_name = 'small_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_small.to_excel(writer, sheet_name = 'small_tests')
        table_step1_med.to_excel(writer, sheet_name = 'med_step1', rename_index = rename_index_step1)
        table_step2_w_med.to_excel(writer, sheet_name = 'med_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_med.to_excel(writer, sheet_name = 'med_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_med.to_excel(writer, sheet_name = 'med_tests')
        table_step1_large.to_excel(writer, sheet_name = 'large_step1', rename_index = rename_index_step1)
        table_step2_w_large.to_excel(writer, sheet_name = 'large_step2_w', rename_index = rename_index_step2_w)
        table_step2_ls_large.to_excel(writer, sheet_name = 'large_step2_ls', rename_index = rename_index_step2_ls)
        df_tests_large.to_excel(writer, sheet_name = 'large_tests')
        
        writer.save()
        writer.close()  