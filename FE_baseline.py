#------------------------------------------
# Fe baseline for first working paper
# Mark van der Plaat
# October 2019 

 # Import packages
import pandas as pd
import numpy as np
import scipy.stats as sps # used to calculated cdf and pdf

import os
os.chdir(r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char')

# Import method that adds a constant to a df
from statsmodels.tools.tools import add_constant

# Import method that can estimate a pooled probit
from statsmodels.discrete.discrete_model import Probit

# Import method for POLS (also does FE)
from linearmodels import PanelOLS

# Import packages for the Sargan-Hausman test
from linearmodels.iv._utility import annihilate
from linearmodels.utility import WaldTestStatistic

import sys # to use the help functions needed
sys.path.insert(1, r'X:\My Documents\PhD\Coding_docs\Help_functions')

from summary3 import summary_col
#--------------------------------------------
# Set parameters 
log = True # If set to False the program estimates the model without logs and with size
change_ls = True # If set to False the program will run a different subset and append it to the excel

#---------------------------------------------- 
#----------------------------------------------
# Prelims
#----------------------------------------------
#----------------------------------------------

#----------------------------------------------
# Load data and add needed variables

# Load df
df = pd.read_csv('df_wp1_clean.csv', index_col = 0)

## Make multi index
df.date = pd.to_datetime(df.date.astype(str).str.strip() + '1230')
df.set_index(['IDRSSD','date'],inplace=True)

## Drop missings on distance
df.dropna(subset = ['distance'], inplace = True)

## Dummy variable for loan sales
if log:
    df['dum_ls'] = np.exp((df.ls_tot > 0) * 1) - 1 #will be taken the log of later
else:
    df['dum_ls'] = (df.ls_tot > 0) * 1    

## Take a subset of variables (only the ones needed)
vars_needed = ['distance','provratio','rwata','net_coffratio_tot_ta',\
               'allowratio_tot_ta','ls_tot_ta','cd_pur_ta','cd_sold_ta',\
               'RC7205','loanratio','roa','depratio','comloanratio','RC2170',\
               'dum_ls','size','bhc']
df_full = df[vars_needed]

## drop NaNs
df_full.dropna(subset = ['provratio','rwata','net_coffratio_tot_ta','allowratio_tot_ta',\
               'ls_tot_ta','cd_pur_ta','cd_sold_ta', 'RC7205','loanratio','roa',\
               'depratio','comloanratio','RC2170','size'], inplace = True)

#---------------------------------------------------
# Setup the data

## Set aside TA
ta = df_full.RC2170

## Take logs of the df
if log:
    df_full = df_full.transform(lambda df: np.log(1 + df)).replace([np.inf, -np.inf], 0)

## Add TA for subsetting
df_full['ta_sub'] = ta

## Add the x_xbar to the df
if log:
    x_xbar = df_full[['cd_pur_ta','cd_sold_ta','RC7205','loanratio','roa',\
                      'depratio','comloanratio','RC2170']].transform(lambda df: df - df.mean())
    df_full[[x + '_xbar' for x in ['cd_pur_ta','cd_sold_ta','RC7205','loanratio',\
                                   'roa','depratio','comloanratio','RC2170']]] = x_xbar
else:
    x_xbar = df_full[['cd_pur_ta','cd_sold_ta','RC7205','loanratio','roa',\
                          'depratio','comloanratio','size']].transform(lambda df: df - df.mean())
    df_full[[x + '_xbar' for x in ['cd_pur_ta','cd_sold_ta','RC7205','loanratio',\
                                   'roa','depratio','comloanratio','size']]] = x_xbar

# Subset the df 
## Only take the banks that change in dum_ls
if change_ls:
    intersect = np.intersect1d(df_full[df_full.ls_tot_ta > 0].index.\
                               get_level_values(0).unique(),\
                               df_full[df_full.ls_tot_ta == 0].index.\
                               get_level_values(0).unique())
    df_sub = df_full[df_full.index.get_level_values(0).isin(intersect)]
else:
    ## Kick out the community banks (based on Stiroh, 2004)  
    ids_comm = df_full[((df_full.index.get_level_values(1) == pd.Timestamp(2018,12,30)) &\
                     (df_full.ta_sub < 3e5) & (df_full.bhc == 0))].index.get_level_values(0).unique().tolist() 
    ids_tot = df_full.index.get_level_values(0).unique().tolist()
    ids_sub = [x for x in ids_tot if x not in ids_comm]
    df_sub = df_full[df_full.index.get_level_values(0).isin(ids_sub)]   

## Add dummies
dummy_full = pd.get_dummies(df_full.index.get_level_values(1))
dummy_sub = pd.get_dummies(df_sub.index.get_level_values(1))

### Add dummies to the dfs
col_dummy = ['dum' + dummy for dummy in dummy_full.columns.astype(str).str[:4].tolist()]
dummy_full = pd.DataFrame(np.array(dummy_full), index = df_full.index, columns = col_dummy)
dummy_sub = pd.DataFrame(np.array(dummy_sub), index = df_sub.index, columns = col_dummy)
df_full[col_dummy] = dummy_full
df_sub[col_dummy] = dummy_sub

## Take the first differences
df_full_fe = df_full.groupby(df_full.index.get_level_values(0)).transform(lambda df: df - df.mean()).dropna()
df_sub_fe = df_sub.groupby(df_sub.index.get_level_values(0)).transform(lambda df: df  - df.mean()).dropna()

#----------------------------------------------
#----------------------------------------------
# Analyses
#----------------------------------------------
#----------------------------------------------

# Set the righthand side of the formulas
if log:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + RC2170'
else:
    righthand_x = r'RC7205 + loanratio + roa + depratio + comloanratio + size'

time_dummies = ' + '.join(col_dummy[1:])

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 1: Full Sample
#----------------------------------------------
'''-----------------------------------------''' 

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_full_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_charge_w.summary)

#----------------------------------------------
## Allowance 
res_full_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_allow_w.summary)

#----------------------------------------------
## rwata 
res_full_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_rwata_w.summary)

#----------------------------------------------
## prov 
res_full_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_full_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_charge_ls.summary)

#----------------------------------------------
## Allowance
res_full_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_allow_ls.summary)

#----------------------------------------------
## rwata
res_full_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_rwata_ls.summary)

#----------------------------------------------
## prov
res_full_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_full_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_full_step2a_prov_ls.summary)

'''-----------------------------------------''' 
#----------------------------------------------
# MODEL 2: Subsample
#----------------------------------------------
'''-----------------------------------------''' 

#----------------------------------------------
# Dummy LS

#----------------------------------------------
## Charge off 
res_sub_step2a_charge_w = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_charge_w.summary)

#----------------------------------------------
## Allowance 
res_sub_step2a_allow_w = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_allow_w.summary)

#----------------------------------------------
## rwata 
res_sub_step2a_rwata_w = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_rwata_w.summary)

#----------------------------------------------
## prov 
res_sub_step2a_prov_w = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_prov_w.summary)

#----------------------------------------------
# LS/TA

#----------------------------------------------
## Charge off 
res_sub_step2a_charge_ls = PanelOLS.from_formula('net_coffratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_charge_ls.summary)

#----------------------------------------------
## Allowance
res_sub_step2a_allow_ls = PanelOLS.from_formula('allowratio_tot_ta' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_allow_ls.summary)

#----------------------------------------------
## rwata
res_sub_step2a_rwata_ls = PanelOLS.from_formula('rwata' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_rwata_ls.summary)

#----------------------------------------------
## prov
res_sub_step2a_prov_ls = PanelOLS.from_formula('provratio' + ' ~ ' + righthand_x + ' + ' + time_dummies, data = df_sub_fe).fit(cov_type = 'clustered', cluster_entity = True)
print(res_sub_step2a_prov_ls.summary)

# Make tables
table_step2_w_full = summary_col([\
    res_full_step2a_charge_w, res_full_step2a_allow_w, res_full_step2a_rwata_w, res_full_step2a_rwata_w, res_full_step2a_prov_w],\
                     show = 'se', regressor_order = ['G_hat_fe_w',\
               'RC7205','loanratio','roa','depratio','comloanratio'])
table_step2_w_sub = summary_col([\
    res_sub_step2a_charge_w,res_sub_step2a_allow_w,res_sub_step2a_rwata_w,res_sub_step2a_prov_w],\
                     show = 'se', regressor_order = ['G_hat_fe_w',\
               'RC7205','loanratio','roa','depratio','comloanratio'])

table_step2_ls_full = summary_col([\
    res_full_step2a_charge_ls,res_full_step2a_allow_ls, res_full_step2a_rwata_ls, res_full_step2a_prov_ls],\
                     show = 'se', regressor_order = ['G_hat_fe_ls',\
               'RC7205','loanratio','roa','depratio','comloanratio'])
table_step2_ls_sub = summary_col([\
    res_sub_step2a_charge_ls,res_sub_step2a_allow_ls,res_sub_step2a_rwata_ls,res_sub_step2a_prov_ls],\
                     show = 'se', regressor_order = ['G_hat_fe_ls',\
               'RC7205','loanratio','roa','depratio','comloanratio'])

#-----------------------------------------------
# Save to a single excel
from openpyxl import load_workbook
path = r'X:\My Documents\PhD\Materials_dissertation\2-Chapter_2-bank_char\FD_baseline_results.xlsx'

if log:
    if change_ls:
        with pd.ExcelWriter('FE_baseline_results.xlsx') as writer:
            table_step2_w_full.to_excel(writer, sheet_name = 'Full_step2_w_log')
            table_step2_ls_full.to_excel(writer, sheet_name = 'Full_step2_ls_log')
            table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_w_log')
            table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_ls_log')
    else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_w_log')
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_ls_log')
        
        writer.save()
        writer.close()  
           
else:
    if change_ls:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        
        table_step2_w_full.to_excel(writer, sheet_name = 'Full_step2_w')
        table_step2_ls_full.to_excel(writer, sheet_name = 'Full_step2_ls')
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_w')
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_changels_step2_ls')
    
        writer.save()
        writer.close()
    else:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book                    
    
        table_step2_w_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_w')
        table_step2_ls_sub.to_excel(writer, sheet_name = 'Sub_nocomm_step2_ls')
        
        writer.save()
        writer.close()  
                       

